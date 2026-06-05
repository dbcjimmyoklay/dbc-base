"""
main.py
DBC 影子中台 — 主控程式
─────────────────────────────────────────────
使用方式：

  每日自動執行（CLEANER → WATCHER → ROOM → REPORTER）：
    python main.py

  只執行指定工作員：
    python main.py --cleaner
    python main.py --watcher
    python main.py --room
    python main.py --reporter

  手動觸發（填完野澤訂房表後執行）：
    python main.py --checkin

  執行全部（含 CHECKIN）：
    python main.py --all

  略過某個工作員：
    python main.py --skip-watcher
    python main.py --skip-room
─────────────────────────────────────────────
"""

import sys
import os
import time
import argparse
import traceback
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)


# ═══════════════════════════════════════════
# 完整大表.xlsx 輸出（本機留存，含個資）
# ═══════════════════════════════════════════

FULL_COLS = [
    '項目', 'OP備註', '出發日期', '泊數', '天數', '註記', '訂編',
    '中文姓名', '英文姓名', '性別', '年齡',
    '尺寸', '身高', '體重', '腳長', '雪板',
    'LEVEL', '類別', '衣褲', '護膝', '護臀', '手套', '4件組',
    '教練', '助教', 'checking', '備註', '報名日期', '團費',
    '飲食', '房號', '序號', '分房備註',
    '手機號碼', 'EMail', '國籍', '身分證', '生日',
    '護照號碼', '護照到期日', '旅客編號',
]


def export_full_xlsx(all_records_flat):
    """輸出完整大表.xlsx（含個資，本機留存，不上雲）"""
    from shared.utils import clean_val

    wb   = Workbook()
    ws   = wb.active
    ws.title = '完整大表'

    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(FULL_COLS, 1):
        cell       = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font

    BLACK_FILL = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    row_num    = 2
    prev_date  = None

    for r in all_records_flat:
        cur_date = r.get('出發日期', '')
        if cur_date != prev_date:
            for col_idx in range(1, len(FULL_COLS) + 1):
                ws.cell(row=row_num, column=col_idx, value='').fill = BLACK_FILL
            row_num  += 1
            prev_date = cur_date
        for col_idx, col_name in enumerate(FULL_COLS, 1):
            ws.cell(row=row_num, column=col_idx, value=clean_val(r.get(col_name, '')))
        row_num += 1

    for i in range(1, len(FULL_COLS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    out_path = os.path.join(BASE_DIR, '完整大表.xlsx')
    wb.save(out_path)
    print(f"  OK 完整大表.xlsx 輸出完成（{len(all_records_flat)} 筆，含個資）")


# ═══════════════════════════════════════════
# 執行單一工作員（含計時與錯誤處理）
# ═══════════════════════════════════════════

def run_agent(name, fn, *args, **kwargs):
    """執行單一工作員，回傳 (success, result, elapsed_sec)"""
    print(f"\n{'-' * 50}")
    print(f">>  {name}")
    print(f"{'-' * 50}")
    t0 = time.time()
    try:
        result = fn(*args, **kwargs)
        elapsed = time.time() - t0
        print(f"OK  {name} 完成（{elapsed:.1f}s）")
        return True, result, elapsed
    except Exception as e:
        elapsed = time.time() - t0
        print(f"NG  {name} 失敗（{elapsed:.1f}s）：{e}")
        traceback.print_exc()
        return False, None, elapsed


# ═══════════════════════════════════════════
# 主流程
# ═══════════════════════════════════════════

def run_checkout():
    """送客單：每日流程自動觸發（COURSE 之後）"""
    from agents.checkout import run as checkout_run
    run_agent('CHECKOUT 送客員', checkout_run)


def run_sales():
    """銷量分析：讀當季+歷年大總表，輸出 portal/sales_data.json"""
    from agents.sales import run as sales_run
    run_agent('SALES 銷量分析員', sales_run)


def run_daily(skip_watcher=False, skip_reporter=False):
    """
    每日自動流程：
    CLEANER → COURSE → CHECKOUT → WATCHER → REPORTER → 完整大表.xlsx
    """
    from agents.cleaner  import run as cleaner_run
    from agents.course   import run as course_run
    from agents.watcher  import run as watcher_run
    from agents.reporter import run as reporter_run

    results   = {}
    t_start   = time.time()
    cleaner_output = None

    # 1. CLEANER（必須成功，後續都依賴它）
    ok, cleaner_output, _ = run_agent('CLEANER 清洗員', cleaner_run)
    results['cleaner'] = ok
    if not ok:
        print("\n[STOP] CLEANER 失敗，中止後續流程")
        return results

    # 2. COURSE（寫入課程安排）
    ok, _, _ = run_agent('COURSE 課程安排寫入員', course_run, cleaner_output)
    results['course'] = ok

    # 3. WATCHER
    if not skip_watcher:
        ok, _, _ = run_agent('WATCHER 比對員', watcher_run, cleaner_output)
        results['watcher'] = ok
    else:
        print("\n  -- 略過 WATCHER --")

    # 5. REPORTER
    if not skip_reporter:
        ok, _, _ = run_agent('REPORTER 報表員', reporter_run, cleaner_output)
        results['reporter'] = ok
    else:
        print("\n  -- 略過 REPORTER --")

    # 6. SALES（銷量分析 JSON，給 Portal 用）
    try:
        from agents.sales import run as sales_run
        ok, _, _ = run_agent('SALES 銷量分析員', sales_run)
        results['sales'] = ok
    except Exception as e:
        print(f"\n  -- SALES 略過：{e} --")
        results['sales'] = False

    # 6. 完整大表.xlsx
    print(f"\n{'-' * 50}")
    print(">>  完整大表.xlsx 輸出")
    print(f"{'-' * 50}")
    try:
        export_full_xlsx(cleaner_output['all_records_flat'])
        results['full_xlsx'] = True
    except Exception as e:
        print(f"NG  完整大表.xlsx 輸出失敗：{e}")
        results['full_xlsx'] = False

    # 總結
    total = time.time() - t_start
    print(f"\n{'=' * 50}")
    print(f"每日流程完成  總耗時 {total:.1f}s  {datetime.now().strftime('%Y/%m/%d %H:%M')}")
    for agent, ok in results.items():
        status = 'OK' if ok else 'NG'
        print(f"  {status} {agent}")
    print(f"{'=' * 50}\n")

    return results


def run_checkin():
    """手動觸發流程（填完野澤訂房表後執行）"""
    from agents.checkin import run as checkin_run
    run_agent('CHECKIN 入住員', checkin_run)


# ═══════════════════════════════════════════
# CLI 入口
# ═══════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='DBC 影子中台主控程式',
        formatter_class=argparse.RawTextHelpFormatter,
    )

    # 單獨執行某個工作員
    parser.add_argument('--cleaner',  action='store_true', help='只執行 CLEANER 清洗員')
    parser.add_argument('--watcher',  action='store_true', help='只執行 WATCHER 比對員')
    # --room 已移除（房務表功能已整併入入住單）
    parser.add_argument('--reporter', action='store_true', help='只執行 REPORTER 報表員')
    parser.add_argument('--checkin',  action='store_true', help='執行 CHECKIN 入住員（手動觸發）')
    parser.add_argument('--checkout', action='store_true', help='只執行 CHECKOUT 送客員')
    parser.add_argument('--sales',    action='store_true', help='只執行 SALES 銷量分析員')
    parser.add_argument('--all',      action='store_true', help='執行全部（含 CHECKIN）')

    # 略過某個工作員
    parser.add_argument('--skip-watcher',  action='store_true', help='每日流程略過 WATCHER')
    parser.add_argument('--skip-reporter', action='store_true', help='每日流程略過 REPORTER')

    args = parser.parse_args()

    print(f"\n{'=' * 50}")
    print(f"DBC 影子中台  {datetime.now().strftime('%Y/%m/%d %H:%M')}")
    print(f"{'=' * 50}")

    # ── 單獨執行 ──
    if args.cleaner:
        from agents.cleaner import run as cleaner_run
        run_agent('CLEANER 清洗員', cleaner_run)
        return

    if args.watcher:
        from agents.watcher import run as watcher_run
        run_agent('WATCHER 比對員', watcher_run)
        return

    if args.reporter:
        from agents.reporter import run as reporter_run
        run_agent('REPORTER 報表員', reporter_run)
        return

    if args.checkin:
        run_checkin()
        return

    if args.checkout:
        run_checkout()
        return

    if args.sales:
        run_sales()
        return

    # ── 全部（含 CHECKIN）──
    if args.all:
        results = run_daily(
            skip_watcher  = args.skip_watcher,
            skip_reporter = args.skip_reporter,
        )
        if results.get('cleaner'):
            run_checkin()
        return

    # ── 預設：每日自動流程 ──
    run_daily(
        skip_watcher  = args.skip_watcher,
        skip_reporter = args.skip_reporter,
    )


if __name__ == '__main__':
    main()
