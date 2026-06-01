"""
DBC 科威大總表清洗腳本 v3.0
輸入：原始大總表*.xlsx
輸出：
  - Google Sheet「課程安排26/27」← 直接寫入，含比對異動
  - Google Sheet「異動紀錄」← 異動記錄
  - 完整大表.xlsx ← 本機留存（含個資，不上雲）
"""

import pandas as pd
import re
import glob
import os
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.service_account import Credentials

# ===== Google Sheets 設定 =====
SPREADSHEET_ID = '12p71jgMErzZYO4toU2LVELDPfwklHCyDARclldYImCc'
SHEET_COURSE   = '課程安排26/27'
SHEET_CHANGE   = '異動紀錄'

folder = os.path.dirname(os.path.abspath(__file__))
CREDS_PATH = os.path.join(folder, 'dbc-credentials.json')

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

# 自動欄位（科威來的，每次覆蓋）
AUTO_COLS = {
    '項目', '出發日期', '泊數', '天數', '註記', '訂編',
    '中文姓名', '英文姓名', '性別', '年齡',
    '尺寸', '身高', '體重', '腳長', '雪板',
    'LEVEL', '類別', '衣褲', '護膝', '護臀', '手套', '4件組'
}

# 手動欄位（人員填的，永不覆蓋）
MANUAL_COLS = {'教練', '助教', 'checking', '備註'}

# 異動要通知的重要欄位
IMPORTANT_COLS = {'Lv', '出發日期', '天數', '類別', '衣褲', '護膝', '護臀', '手套', '4件組', '項目', '泊數'}

# 課程安排表欄位順序
COURSE_COLS = [
    '編號', '項目', '出發日期', '泊數', '天數', '註記', '訂編',
    '中文姓名', '英文姓名', '性別', '年齡',
    '尺寸', '身高', '體重', '腳長', '雪板',
    'LEVEL', '類別', '衣褲', '護膝', '護臀', '手套', '4件組',
    '教練', '助教', 'checking', '備註',
    '旅客編號', '項目1'  # 隱藏欄位
]

# 表頭顯示名稱
COURSE_HEADERS = [
    '編號', '項目', '出發日期', '泊數', '天數', '註記', '訂編',
    '中文姓名', '英文姓名', '性別', '年齡',
    '尺寸', '身高', '體重', '腳長', '雪板',
    'Lv', '類別', '衣褲', '護膝', '護臀', '手套', '四件',
    '教練', '助教', 'checking', '備註',
    '旅客編號', '項目1'  # 隱藏欄位（人員不需要看）
]

# 隱藏欄位索引（寫入後自動隱藏）
HIDDEN_COLS = {'旅客編號', '項目1'}

# 房務表設定（湯澤純課不寫入）
SHEET_ROOM = '房務表26/27'
ROOM_EXCLUDE_ITEMS = {'湯澤純課'}
ROOM_COLS = [
    '項目', '出發日期', '泊數', '訂編',
    '中文姓名', '英文姓名', '性別', '年齡',
    '房號', '序號', '備註', '飲食', 'OP備註'
]
ROOM_HEADERS = ROOM_COLS
ROOM_MANUAL_COLS = {'房號', '序號', '備註'}
ROOM_IMPORTANT_COLS = {'出發日期', '泊數', '項目', '飲食'}
ROOM_CENTER_COLS = {
    '項目', '出發日期', '泊數', '訂編',
    '中文姓名', '英文姓名', '性別', '年齡',
    '房號', '序號'
}

# ===== 入住單設定 =====
CHECKIN_COLS = [
    '項目', '出發日期', '泊數', '天數', '訂編',
    '中文姓名', '英文姓名', '性別', '年齡',
    '房型', '入住備註'
]
CHECKIN_HEADERS = CHECKIN_COLS
CHECKIN_MANUAL_COLS = {'房型', '入住備註'}
CHECKIN_IMPORTANT_COLS = {'出發日期', '泊數', '天數', '項目'}
CHECKIN_CENTER_COLS = {
    '項目', '出發日期', '泊數', '天數', '訂編',
    '中文姓名', '英文姓名', '性別', '年齡', '房型'
}

# 各區入住單：分頁名稱 / gid / 只取哪些項目
CHECKIN_SHEETS = [
    {
        'name': '野澤入住單26/27',
        'gid': 2070775881,
        'include': lambda item: '野澤' in item and '純課' not in item,
    },
    {
        'name': '斑尾入住單26/27',
        'gid': 460046160,
        'include': lambda item: '斑尾' in item and '純課' not in item,
    },
    {
        'name': '龍平入住單26/27',
        'gid': 343688459,
        'include': lambda item: '龍平' in item,
    },
]

# ===== 雪場對應規則 =====
ITEM_RULES = [
    ('ZN', '野澤純課'), ('ZB', '斑尾純課'), ('ZY', '湯澤純課'),
    ('YS', '龍平'), ('NT', '野澤'), ('BT', '斑尾'),
    ('N', '野澤'), ('B', '斑尾'),
]
PURE_ITEMS = {'野澤純課', '斑尾純課', '湯澤純課'}
RESORT_NAME = {'野澤純課': '野澤', '斑尾純課': '斑尾', '湯澤純課': '湯澤'}
ITEM_ORDER = {'野澤':1,'野澤純課':2,'斑尾':3,'斑尾純課':4,'湯澤純課':5,'龍平':6}

# ===== 價目表 =====
PRICE_TABLE = {
    '野澤純課': {1:16500,2:17500,3:19500,4:19500,5:21000,6:21000},
    '斑尾純課': {1:14500,2:15000,3:16000,4:17000,5:18000,6:19000},
    '湯澤純課': {1:12500,2:13000,3:14000,4:14000,5:16000,6:16000},
}
YUKIGUNI_PEAK = [
    (1,12,1,15,1000),(1,21,2,10,1000),(2,25,2,25,1000),
    (2,26,3,2,2000),(3,3,3,7,1000),
]
EQUIP_PRICE = {'衣褲':1000,'護膝':200,'護臀':200,'4件組':1800,'手套':400}

WEEKDAY_MAP = {0:'一',1:'二',2:'三',3:'四',4:'五',5:'六',6:'日'}

# ===== 工具函數 =====

def get_weekday(date):
    return WEEKDAY_MAP.get(date.weekday(), '') if date else ''

def get_yukiguni_surcharge(lesson_date, year):
    if not lesson_date: return 0
    for (m1,d1,m2,d2,s) in YUKIGUNI_PEAK:
        y1 = year if m1>=12 else year+1
        y2 = year if m2>=12 else year+1
        if datetime(y1,m1,d1) <= lesson_date <= datetime(y2,m2,d2):
            return s
    return 0

def calc_person_equip_fee(has_equip, days):
    fee = 0
    for key, price in EQUIP_PRICE.items():
        if has_equip.get(key, False):
            fee += price if key == '手套' else price * days
    return fee

def infer_days_from_fee(item, order_members, lesson_date=None, season_start_year=2026):
    if item not in PRICE_TABLE: return None, None
    n_people = len(order_members)
    if n_people < 1 or n_people > 6: return None, None
    surcharge = get_yukiguni_surcharge(lesson_date, season_start_year) if item == '湯澤純課' and lesson_date else 0
    for ratio in range(1,7):
        base_price = PRICE_TABLE[item].get(ratio,0)
        day_price = base_price + surcharge
        if not day_price: continue
        for days in range(1,8):
            total = sum(m['fee'] - calc_person_equip_fee(m['has_equip'],days) for m in order_members)
            if abs(total - day_price * days) < 1:
                return days, ratio
    return None, None

def cn_to_num(s):
    for k,v in {'一':'1','二':'2','三':'3','四':'4','五':'5','六':'6','七':'7','八':'8','九':'9','零':'0'}.items():
        s = s.replace(k,v)
    return s

def get_item(g):
    if pd.isna(g): return '其他'
    s = str(g)
    for key, label in ITEM_RULES:
        if key in s: return label
    return '其他'

def extract_date_from_group(g):
    if pd.isna(g): return None
    m = re.search(r'(\d{6})', str(g))
    if m:
        d = m.group(1)
        try: return datetime(2000+int(d[:2]),int(d[2:4]),int(d[4:6]))
        except: return None
    return None

def has_value(row, col):
    if col not in row.index: return False
    v = row[col]
    return pd.notna(v) and v != 0 and v != ''

def safe_int(val):
    try:
        if val is None: return ''
        try:
            if pd.isna(val): return ''
        except: pass
        v = pd.to_numeric(val, errors='coerce')
        if pd.isna(v): return ''
        i = int(v)
        return '' if i == 0 else i  # 0值也回傳空字串
    except: return ''

def clean_order_no(val):
    try: return str(int(str(val).lstrip('0')))
    except: return str(val)

def parse_op_section(op, order_no):
    if not op or pd.isna(op): return ''
    try: ono = str(int(str(order_no).lstrip('0')))
    except: ono = str(order_no)
    text = cn_to_num(str(op))
    m = re.search(rf'#\s*0*{re.escape(ono)}\b[\s\S]*?(?=\n#\d|\Z)', text)
    return m.group() if m else text

def parse_op_days(op, order_no):
    s = parse_op_section(op, order_no)
    for pat in [r'共\s*(\d+)\s*[日天晚泊]',r'上課\s*(\d+)\s*[日天]',r'(\d+)\s*[日天]\s*課程']:
        m = re.search(pat, s)
        if m: return int(m.group(1))
    return None

def parse_op_ratio(op, order_no):
    s = parse_op_section(op, order_no)
    m = re.search(r'1\s*[對对:]\s*(\d+)', s)
    return f"1對{m.group(1)}" if m else None

def clean_diet(raw):
    """清洗飲食備註欄位，只保留有意義的標籤"""
    if raw is None: return ''
    if hasattr(raw, 'iloc'):
        raw = raw.iloc[0] if len(raw) > 0 else ''
    try:
        if pd.isna(raw): return ''
    except: pass
    s = str(raw).strip()
    if not s or s.lower() in ('nan', 'none'): return ''

    CODE_MAP = {
        'VLML': '奶蛋素',
        'NBML': '非牛肉餐',
        'AVML': '東方素食',
        'CHML': '兒童餐',
        'GFML': '無麩質',
        'LCML': '低卡',
        'LSML': '低鹽',
        'NLML': '無乳糖',
    }
    SKIP_CODES = {'NAML'}

    # 抽出所有完整代碼（4個大寫字母）
    codes = re.findall(r'\b([A-Z]{4})\b', s)
    results = []
    seen = set()
    for code in codes:
        if code in SKIP_CODES: continue
        label = CODE_MAP.get(code)
        if label and label not in seen:
            results.append(label)
            seen.add(label)

    # 只有在沒有抓到任何代碼時，才嘗試保留自由文字
    if not results:
        # 移除所有代碼相關文字，保留純自由文字
        free = re.sub(r'[A-Z]{4}[_\s－/／]*[^\s,，A-Z]*', '', s).strip().strip(',，／/').strip()
        if free and free not in ('', '無指定'):
            results.append(free)

    return '／'.join(results) if results else ''

def clean_val(x):
    if x is None: return ''
    if hasattr(x, 'iloc'):
        x = x.iloc[0] if len(x) > 0 else ''
    try:
        if pd.isna(x): return ''
    except: pass
    s = str(x).strip()
    if s.lower() in ('nan','none',''): return ''
    if s in ('0','0.0'): return ''
    if s.endswith('.0') and s[:-2].lstrip('-').isdigit(): return s[:-2]
    return s

def classify_cols(cols):
    pure, excl, with_, without_, four = {}, [], {}, {}, []
    rent = {'衣褲':[],'護膝':[],'護臀':[],'手套':[]}
    keep_city = {}  # 保留市區住宿：天數+N，泊數不變
    for col in cols:
        c = str(col)
        if '純課程' in c:
            m = re.search(r'1[對对]\s*(\d+)', c)
            if m: pure[col] = int(m.group(1))
        if '專屬教練' in c: excl.append(col)
        if ('續滑' in c or '加滑' in c) and '含教練' in c and '不含教練' not in c and '保留市區住宿' not in c:
            m = re.search(r'(\d+)\s*[日天]', c)
            if m: with_[col] = int(m.group(1))
        if ('續滑' in c or '加滑' in c) and '不含教練' in c and '保留市區住宿' not in c:
            m = re.search(r'(\d+)\s*[日天]', c)
            if m: without_[col] = int(m.group(1))
        # 保留市區住宿：天數+N，泊數不增加
        if '保留市區住宿' in c:
            m = re.search(r'(\d+)\s*[日天]', c)
            days = int(m.group(1)) if m else 1  # 沒有數字預設1日
            keep_city[col] = days
        if '4件組' in c or '四件組' in c: four.append(col)
        if '租借' in c:
            if '雪衣' in c: rent['衣褲'].append(col)
            if '護膝' in c: rent['護膝'].append(col)
            if '護臀' in c: rent['護臀'].append(col)
            if '手套' in c: rent['手套'].append(col)
    return pure, excl, with_, without_, four, rent, keep_city

# ===== 讀取科威大總表 =====
files = glob.glob(os.path.join(folder, '原始大總表*.xlsx'))
if not files:
    raise FileNotFoundError("找不到原始大總表*.xlsx，請確認檔案在同一資料夾")
latest_file = max(files, key=os.path.getmtime)
print(f"使用檔案：{os.path.basename(latest_file)}")

df_raw = pd.read_excel(latest_file, header=None)
headers = df_raw.iloc[1].tolist()

# 重複欄位改名（序號→序號_1/序號_2，分房備住→分房備住_1/分房備住_2）
seen_headers = {}
new_headers = []
for h in headers:
    h_str = str(h)
    if h_str in seen_headers:
        seen_headers[h_str] += 1
        new_headers.append(f"{h_str}_{seen_headers[h_str]}")
    else:
        seen_headers[h_str] = 1
        new_headers.append(h_str)

df = df_raw.iloc[2:].copy()
df.columns = new_headers
df = df.reset_index(drop=True)
df = df[df['團號'] != '數量小計'].copy()
df['團號'] = df['團號'].ffill()
df = df[df['訂單編號'].notna()].copy()
df = df.reset_index(drop=True)

COL_SEQ = 0   # 不再需要位置索引
COL_FENBEI = 0
print(f"欄位重命名完成，序號→序號_1/序號_2，分房備住→分房備住_1/分房備住_2")

pure_cols, excl_cols, with_cols, without_cols, four_cols, rent_cols, keep_city_cols = classify_cols(df.columns)

if keep_city_cols:
    print(f"偵測到保留市區住宿欄位：{list(keep_city_cols.keys())}")

# 雪季偵測
try:
    dates = pd.to_datetime(df['出發日期'], errors='coerce').dropna()
    y = dates.min().year if dates.min().month >= 10 else dates.min().year - 1
    DATE_START = datetime(y, 12, 5)
    DATE_END = datetime(y+1, 3, 31)
except:
    DATE_START = datetime(2026, 12, 5)
    DATE_END = datetime(2027, 3, 31)

print(f"雪季：{DATE_START.strftime('%Y/%m/%d')} ~ {DATE_END.strftime('%Y/%m/%d')}")
print(f"總資料筆數：{len(df)}")

# ===== 純課程：OP解析 + 團費反推 =====
order_op_cache = {}
for order_no, group in df.groupby('訂單編號', sort=False):
    first_row = group.iloc[0]
    item = get_item(first_row.get('團號',''))
    if item not in PURE_ITEMS: continue
    op = first_row.get('OP備註','')
    days_val = parse_op_days(op, order_no)
    ratio_val = parse_op_ratio(op, order_no)
    tiansu = str(days_val) if days_val else ''
    note = ratio_val or ''
    # 判斷是否專屬教練
    is_excl = any(has_value(first_row, c) for c in excl_cols)
    if is_excl:
        note = '專屬'
    elif not tiansu:
        members = []
        for _, mrow in group.iterrows():
            fee = safe_int(mrow.get('團費',0))
            if fee == '': fee = 0
            has_equip = {
                '衣褲': any(has_value(mrow,c) for c in rent_cols['衣褲']),
                '護膝': any(has_value(mrow,c) for c in rent_cols['護膝']),
                '護臀': any(has_value(mrow,c) for c in rent_cols['護臀']),
                '手套': any(has_value(mrow,c) for c in rent_cols['手套']),
                '4件組': any(has_value(mrow,c) for c in four_cols),
            }
            members.append({'fee':fee,'has_equip':has_equip})
        group_date = extract_date_from_group(first_row.get('團號',''))
        inferred_days, n_people = infer_days_from_fee(item, members, lesson_date=group_date, season_start_year=DATE_START.year)
        if inferred_days:
            tiansu = str(inferred_days)
            if not note and n_people:
                note = f"1對{n_people}"
        else:
            tiansu = ''
            if not note: note = '待確認'
    order_op_cache[str(order_no)] = (tiansu, note)

# ===== 建立完整記錄 =====
records = []
for _, row in df.iterrows():
    group_no = row.get('團號','')
    item = get_item(group_no)
    is_pure = item in PURE_ITEMS

    group_date = extract_date_from_group(group_no)
    if group_date is None:
        try: group_date = pd.to_datetime(row['出發日期'])
        except: group_date = None
    try: group_date = None if pd.isna(group_date) else group_date
    except: pass

    depart_date = (group_date - timedelta(days=1)) if (is_pure and group_date) else group_date
    sort_date = depart_date if depart_date else datetime(2099,1,1)
    depart_display = depart_date.strftime('%Y/%m/%d') if depart_date else ''

    order_no = row.get('訂單編號','')
    op = row.get('OP備註','')

    if is_pure:
        pou_su = ''
        tiansu, note = order_op_cache.get(str(order_no),('',''))
    else:
        is_excl = any(has_value(row,c) for c in excl_cols)
        add_with    = sum(n for col,n in with_cols.items() if has_value(row,col))
        add_without = sum(n for col,n in without_cols.items() if has_value(row,col))
        add_keep    = sum(n for col,n in keep_city_cols.items() if has_value(row,col))
        # 泊數：保留市區住宿不增加泊數
        pou_su = 3 + add_with + add_without
        # 天數：三種續滑都要計入
        total_days = 2 + add_with + add_without + add_keep
        has_含 = add_with > 0
        has_無 = (add_without + add_keep) > 0
        if has_含 and has_無: tiansu = f"{total_days}含無"
        elif has_含:          tiansu = f"{total_days}含"
        elif has_無:          tiansu = f"{total_days}無"
        else:                 tiansu = str(total_days)
        note = '專屬' if is_excl else ''

    equip_yi   = '1' if any(has_value(row,c) for c in rent_cols['衣褲']) else ''
    equip_xi   = '1' if any(has_value(row,c) for c in rent_cols['護膝']) else ''
    equip_tun  = '1' if any(has_value(row,c) for c in rent_cols['護臀']) else ''
    equip_tao  = '1' if any(has_value(row,c) for c in rent_cols['手套']) else ''
    equip_four = '1' if any(has_value(row,c) for c in four_cols) else ''

    lv = clean_val(row.get('LEVEL',''))
    lei = clean_val(row.get('滑雪類別',''))

    records.append({
        '_sort_date':  sort_date,
        '_item_order': ITEM_ORDER.get(item,7),
        '_depart_date': depart_date,
        '項目':    item,
        '出發日期': depart_display,
        '泊數':    str(pou_su) if pou_su != '' else '',
        '天數':    tiansu,
        '註記':    note,
        '訂編':    clean_order_no(order_no),
        '中文姓名': clean_val(row.get('中文姓名','')),
        '英文姓名': clean_val(row.get('英文姓名','')),
        '性別':    clean_val(row.get('性別','')),
        '年齡':    str(safe_int(row.get('年齡'))),
        '尺寸':    '',
        '身高':    str(safe_int(row.get('身高(CM)'))),
        '體重':    str(safe_int(row.get('體重(KG)'))),
        '腳長':    clean_val(row.get('腳長','')),
        '雪板':    '',
        'LEVEL':   lv,
        '類別':    lei,
        '衣褲':    equip_yi,
        '護膝':    equip_xi,
        '護臀':    equip_tun,
        '手套':    equip_tao,
        '4件組':   equip_four,
        '教練':    '',
        '助教':    '',
        'checking': '',
        '備註':    '',
        '新增/刪減/異動': '',
        '旅客編號': clean_val(row.get('旅客編號', '')),  # 隱藏key欄
        '項目1':   item,  # 隱藏原始項目欄（用來判斷是否手動改過）
        # 個資（只存本機）
        'OP備註':  clean_val(row.get('OP備註','')),
        '報名日期': clean_val(row.get('報名日期','')),
        '團費':    clean_val(row.get('團費','')),
        '手機號碼': clean_val(row.get('手機號碼','')),
        'EMail':   clean_val(row.get('EMail','')),
        '國籍':    clean_val(row.get('國籍','')),
        '身分證':  clean_val(row.get('身分證','')),
        '生日':    clean_val(row.get('生日','')),
        '護照號碼': clean_val(row.get('護照號碼','')),
        '護照到期日': clean_val(row.get('護照到期日','')),
        '旅客編號': clean_val(row.get('旅客編號','')),
        '飲食':    clean_diet(row.get('飲食備註\n(定位備註)','')),
        '房號':    clean_val(row.get('標準分房(房號)','')),
        '序號':    clean_val(row.get('序號','')),
        '_房務備註': clean_val(row.get('分房備住','')),
    })

# ===== 排序 =====
records.sort(key=lambda x: (x['_sort_date'], x['_item_order'], x['訂編']))

# ===== 按日期分組，建立每日統計 =====
date_map = defaultdict(list)
for r in records:
    sd = r['_sort_date']
    if DATE_START <= sd <= DATE_END:
        date_map[sd].append(r)

# ===== 建立寫入Google Sheet的資料（含每日分隔列）=====
def normalize_level(lv):
    """統一LEVEL格式為lv0~lv4，空的回傳空字串"""
    if not lv or str(lv).strip() in ['', 'nan', 'None']: return ''
    s = str(lv).strip().lower()
    # 抽出數字
    m = re.search(r'(\d+)', s)
    if m:
        n = int(m.group(1))
        return f"lv{n}"
    return ''

def count_by_area(recs):
    counts = {'野澤':0,'斑尾':0,'湯澤':0,'龍平':0}
    for r in recs:
        item = r.get('項目','')
        if '野澤' in item: counts['野澤'] += 1
        elif '斑尾' in item: counts['斑尾'] += 1
        elif '湯澤' in item: counts['湯澤'] += 1
        elif '龍平' in item: counts['龍平'] += 1
    return counts

all_sheet_rows = []   # 最終寫入Google Sheet的列
all_records_flat = [] # 純資料記錄（不含分隔列，供比對用）
seq_counter = 0

# 產生雪季所有日期（12/05 ~ 03/31）
all_season_dates = []
cur = DATE_START
while cur <= DATE_END:
    all_season_dates.append(cur)
    cur += timedelta(days=1)

for sd in all_season_dates:
    day_records = date_map.get(sd, [])  # 沒有資料的日期回傳空list
    wd = get_weekday(sd)
    date_str = sd.strftime('%y/%m/%d')
    counts = count_by_area(day_records)

    # 分隔列：第一欄空、日期第二欄、曜日第三欄、統計第四至七欄
    sep_row = [''] * len(COURSE_COLS)
    sep_row[1] = date_str
    sep_row[2] = wd
    sep_row[3] = f"野:{counts['野澤']}"
    sep_row[4] = f"斑:{counts['斑尾']}"
    sep_row[5] = f"湯:{counts['湯澤']}"
    sep_row[6] = f"龍:{counts['龍平']}"
    all_sheet_rows.append({'_is_separator': True, '_data': sep_row})

    for local_idx, r in enumerate(day_records, 1):
        r['編號'] = str(local_idx)
        r['出發日期'] = date_str
        r['LEVEL'] = normalize_level(r.get('LEVEL',''))
        all_records_flat.append(r)
        all_sheet_rows.append({'_is_separator': False, '_data': r})

    # 沒有資料的日期加一列空列
    if not day_records:
        empty_row = [''] * len(COURSE_COLS)
        all_sheet_rows.append({'_is_separator': False, '_data': {col: '' for col in COURSE_COLS}})

print(f"整理完成，共 {len(all_records_flat)} 筆資料，{len(date_map)} 個日期")

# ===== 連線 Google Sheets =====
print("連線 Google Sheets...")
creds = Credentials.from_service_account_file(CREDS_PATH, scopes=SCOPES)
gc = gspread.authorize(creds)
sh = gc.open_by_key(SPREADSHEET_ID)

def rebuild_filter(ws, total_data_rows, n_cols, start_row=1):
    """重建分頁篩選器，start_row=1表示從第2列開始（預設），=0表示從第1列開始"""
    sid = ws.id
    last_row = total_data_rows + start_row + 1
    sh.batch_update({'requests': [{'clearBasicFilter': {'sheetId': sid}}]})
    sh.batch_update({'requests': [{
        'setBasicFilter': {
            'filter': {
                'range': {
                    'sheetId': sid,
                    'startRowIndex': start_row,
                    'endRowIndex': last_row,
                    'startColumnIndex': 0,
                    'endColumnIndex': n_cols
                }
            }
        }
    }]})
    print(f"  OK 篩選器重建完成（第{start_row+1}列～第{last_row}列）")

# 確保課程安排分頁存在
try:
    ws_course = sh.worksheet(SHEET_COURSE)
except gspread.exceptions.WorksheetNotFound:
    ws_course = sh.add_worksheet(title=SHEET_COURSE, rows=2000, cols=30)
    print(f"建立新分頁：{SHEET_COURSE}")

# 確保異動紀錄分頁存在
try:
    ws_change = sh.worksheet(SHEET_CHANGE)
except gspread.exceptions.WorksheetNotFound:
    ws_change = sh.add_worksheet(title=SHEET_CHANGE, rows=2000, cols=12)
    # 建立標題列
    ws_change.append_row(['異動日期','異動類型','訂編','中文姓名','異動欄位','原值','新值','處理'])
    print(f"建立新分頁：{SHEET_CHANGE}")

# ===== 讀取現有Google Sheet資料 =====
print("讀取現有課程安排資料...")
existing_data = ws_course.get_all_values()

def build_record_key(r, seq=None):
    """建立唯一key，優先順序：旅客編號 > 訂編+中文姓名 > 訂編+英文姓名 > 訂編+序號"""
    visitor_id = str(r.get('旅客編號', '')).strip()
    if visitor_id:
        return f"v_{visitor_id}"
    ding = str(r.get('訂編', '')).strip()
    cn = str(r.get('中文姓名', '')).strip()
    en = str(r.get('英文姓名', '')).strip()
    if cn:
        return f"{ding}_{cn}"
    if en:
        return f"{ding}_en_{en}"
    return f"{ding}_seq_{seq or 0}"

# 建立現有資料索引：用旅客編號當key（唯一ID）
# 第1列是篩選按鈕，第2列才是標題
existing_map = {}

if len(existing_data) > 1:
    header_row = existing_data[1]  # 第2列是標題
    col_map = {h: idx for idx, h in enumerate(header_row)}
    idx_visitor = col_map.get('旅客編號', -1)
    idx_ding = col_map.get('訂編', -1)
    idx_name = col_map.get('中文姓名', -1)

    for row in existing_data[2:]:
        if not row or len(row) < 2: continue
        visitor_id = row[idx_visitor].strip() if idx_visitor >= 0 and idx_visitor < len(row) else ''
        ding = row[idx_ding].strip() if idx_ding >= 0 and idx_ding < len(row) else ''
        name = row[idx_name].strip() if idx_name >= 0 and idx_name < len(row) else ''
        en_idx = col_map.get('英文姓名', -1)
        en = row[en_idx].strip() if en_idx >= 0 and en_idx < len(row) else ''

        if visitor_id:
            key = f"v_{visitor_id}"
        elif name:
            key = f"{ding}_{name}"
        elif en:
            key = f"{ding}_en_{en}"
        else:
            continue  # 中英文都空且沒旅客編號，跳過

        fields = {}
        for h, idx in col_map.items():
            if idx < len(row):
                fields[h] = row[idx]
        existing_map[key] = fields

print(f"現有資料：{len(existing_map)} 筆")

# ===== 保護機制：讀取異常時停止 =====
if len(existing_data) > 10 and len(existing_map) == 0:
    # 表格有資料但一筆都沒讀到，可能是格式問題
    print("警告：Google Sheet 有資料但讀取失敗，停止執行以保護手動欄位！")
    print("請檢查表格第一列是否為正確標題列")
    input("按Enter結束")
    exit(1)

if len(existing_map) > 0 and len(all_records_flat) > 0:
    ratio = len(existing_map) / len(all_records_flat)
    if ratio < 0.3:
        # 現有資料比新資料少太多，可能讀取有問題
        print(f"警告：現有資料（{len(existing_map)}筆）遠少於新資料（{len(all_records_flat)}筆），可能讀取異常")
        confirm = input("是否繼續執行？手動欄位可能無法保留 (y/n): ")
        if confirm.lower() != 'y':
            print("取消執行")
            exit(0)

# ===== 比對異動 =====
today_str = datetime.now().strftime('%m/%d')
change_logs = []
new_keys = set()

seq_counter_map = {}
for r in all_records_flat:
    ding = r.get('訂編', '')
    seq_counter_map[ding] = seq_counter_map.get(ding, 0) + 1
    r['_seq'] = seq_counter_map[ding]
    key = build_record_key(r, seq=r['_seq'])
    new_keys.add(key)
    item = r.get('項目', '')
    depart = r.get('出發日期', '')

    if key in existing_map:
        old_fields = existing_map[key]
        changes = []
        for col in AUTO_COLS:
            # LEVEL在Google Sheet顯示為Lv，需對應
            sheet_col = 'Lv' if col == 'LEVEL' else col
            old_val = normalize_level(old_fields.get(sheet_col, '')) if col == 'LEVEL' else old_fields.get(sheet_col, '').strip()
            new_val = str(r.get(col, '')).strip()
            if old_val != new_val and col in IMPORTANT_COLS:
                changes.append({'欄位': col, '原值': old_val, '新值': new_val})
        for ch in changes:
            change_logs.append([
                today_str, '異動', item, depart, r['訂編'], r['中文姓名'],
                ch['欄位'], ch['原值'], ch['新值'], ''
            ])
    else:
        change_logs.append([
            today_str, '新增', item, depart, r['訂編'], r['中文姓名'],
            '', '', '', ''
        ])

# 刪減
for key in existing_map:
    if key not in new_keys:
        parts = key.split('_', 1)
        old_item = existing_map[key].get('項目', '')
        old_depart = existing_map[key].get('出發日期', '')
        change_logs.append([
            today_str, '刪減', old_item, old_depart, parts[0], parts[1] if len(parts) > 1 else '',
            '', '', '', ''
        ])

print(f"異動記錄：{len(change_logs)} 筆")

# ===== 寫入異動紀錄 =====
CHANGE_HEADER = ['異動日期', '異動類型', '項目', '出發日期', '訂編', '中文姓名', '異動欄位', '原值', '新值', '處理']

# 先清除已處理（TRUE）的記錄
print("清除已處理的異動紀錄...")
existing_change = ws_change.get_all_values()
if existing_change and len(existing_change) > 1:
    # 第2列是標題（index=1）
    header_row = existing_change[1] if len(existing_change) > 1 else []
    idx_done_col = None
    if header_row:
        try:
            idx_done_col = header_row.index('處理')
        except ValueError:
            pass

    if idx_done_col is not None:
        # 第3列起是資料（index=2）
        kept_rows = []
        removed = 0
        for row in existing_change[2:]:
            done_val = row[idx_done_col].strip().upper() if idx_done_col < len(row) else ''
            if done_val in ('TRUE', '1'):
                removed += 1
            else:
                kept_rows.append(row)
        if removed > 0:
            # 清除第3列以後，重新寫入未處理的資料
            sh.batch_update({'requests': [{
                'updateCells': {
                    'range': {'sheetId': ws_change.id, 'startRowIndex': 2, 'endRowIndex': 2000,
                              'startColumnIndex': 0, 'endColumnIndex': len(CHANGE_HEADER)},
                    'fields': 'userEnteredValue'
                }
            }]})
            if kept_rows:
                ws_change.update(kept_rows, 'A3', value_input_option='RAW')
            existing_change = existing_change[:2] + kept_rows
            print(f"  已清除 {removed} 筆已處理記錄")
        else:
            print("  無已處理記錄需清除")

if change_logs:
    existing_change = ws_change.get_all_values()
    # 不動第1列（篩選按鈕）和第2列（標題+篩選器）
    # 直接從第3列起追加資料
    if len(existing_change) >= 2:
        next_row = len(existing_change) + 1
    else:
        next_row = 3  # 第3列起
    start_row = next_row

    ws_change.update(change_logs, f'A{next_row}', value_input_option='RAW')
    print(f"OK 異動紀錄寫入 {len(change_logs)} 筆")

    # ===== 異動紀錄填色 =====
    print("設定異動紀錄填色...")
    change_sid = ws_change.id

    # 標題列格式不動（由Apps Script管理）
    fmt_reqs = []  # 只處理資料列的格式

    # 欄寬設定（對應 CHANGE_HEADER 10欄）
    # 異動日期/異動類型/項目/出發日期/訂編/中文姓名/異動欄位/原值/新值/處理
    CHANGE_COL_WIDTHS = [65, 65, 65, 70, 55, 80, 80, 80, 80, 65]
    for ci, w in enumerate(CHANGE_COL_WIDTHS):
        fmt_reqs.append({
            'updateDimensionProperties': {
                'range': {'sheetId': change_sid, 'dimension': 'COLUMNS', 'startIndex': ci, 'endIndex': ci+1},
                'properties': {'pixelSize': w}, 'fields': 'pixelSize'
            }
        })

    # 每列填色
    COLOR_NEW    = {'red': 1.0,  'green': 0.95, 'blue': 0.6}   # 亮黃（新增）
    COLOR_DEL    = {'red': 1.0,  'green': 0.7,  'blue': 0.7}   # 紅（刪減）
    COLOR_CHANGE = {'red': 1.0,  'green': 1.0,  'blue': 1.0}   # 白（資料異動）

    idx_type = CHANGE_HEADER.index('異動類型')
    for i, log in enumerate(change_logs):
        ridx = start_row - 1 + i  # 0-based
        t = log[idx_type]
        if t == '新增':      bg = COLOR_NEW
        elif t == '刪減':    bg = COLOR_DEL
        else:                bg = COLOR_CHANGE

        fmt_reqs.append({'repeatCell': {
            'range': {'sheetId': change_sid, 'startRowIndex': ridx, 'endRowIndex': ridx+1,
                      'startColumnIndex': 0, 'endColumnIndex': len(CHANGE_HEADER)},
            'cell': {'userEnteredFormat': {'backgroundColor': bg}},
            'fields': 'userEnteredFormat.backgroundColor'
        }})

    # 格線
    total_change_rows = start_row - 1 + len(change_logs)
    fmt_reqs.append({'updateBorders': {
        'range': {'sheetId': change_sid, 'startRowIndex': 2, 'endRowIndex': total_change_rows,
                  'startColumnIndex': 0, 'endColumnIndex': len(CHANGE_HEADER)},
        'top':    {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'bottom': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'left':   {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'right':  {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'innerHorizontal': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0.8,'green':0.8,'blue':0.8}}},
        'innerVertical':   {'style':'SOLID','colorStyle':{'rgbColor':{'red':0.8,'green':0.8,'blue':0.8}}},
    }})

    # ===== 處理欄設為 Checkbox =====
    idx_done = CHANGE_HEADER.index('處理')
    fmt_reqs.append({'repeatCell': {
        'range': {
            'sheetId': change_sid,
            'startRowIndex': 2,  # 從第3列開始（跳過篩選按鈕+標題）
            'endRowIndex': total_change_rows + 200,
            'startColumnIndex': idx_done,
            'endColumnIndex': idx_done + 1
        },
        'cell': {
            'dataValidation': {
                'condition': {'type': 'BOOLEAN'},
                'strict': True,
                'showCustomUi': True
            }
        },
        'fields': 'dataValidation'
    }})

    # 凍結標題列
    # 不動凍結設定（由使用者自行設定）

    for i in range(0, len(fmt_reqs), 500):
        sh.batch_update({'requests': fmt_reqs[i:i+500]})

    print("OK 異動紀錄格式設定完成")
    print("重建異動紀錄篩選器...")
    rebuild_filter(ws_change, total_change_rows - 2, len(CHANGE_HEADER))

# ===== 寫入課程安排（全部重寫）=====
print("寫入課程安排...")

# 建立要寫入的所有列
write_rows = [COURSE_HEADERS]  # 第一列是顯示用標題

for item in all_sheet_rows:
    if item['_is_separator']:
        write_rows.append(item['_data'])
    else:
        r = item['_data']
        key = build_record_key(r)

        # 保留手動欄位
        if key in existing_map:
            old = existing_map[key]
            for mc in MANUAL_COLS:
                if mc in old and old[mc]:
                    r[mc] = old[mc]
            # 保護「項目」欄：若現有值跟項目1不同，代表手動改過旅館名稱 → 保留
            old_item = old.get('項目', '').strip()
            old_item1 = old.get('項目1', '').strip()
            if old_item and old_item != old_item1:
                r['項目'] = old_item  # 保留旅館名稱

        row_data = []
        for col in COURSE_COLS:
            v = r.get(col, '')
            row_data.append(str(v) if v is not None else '')
        write_rows.append(row_data)

# 清除第2列以後的內容和格式（保留第1列篩選按鈕）
sh.batch_update({'requests': [{
    'updateCells': {
        'range': {
            'sheetId': ws_course.id,
            'startRowIndex': 1,  # 從第2列開始清除
            'endRowIndex': 2000,
            'startColumnIndex': 0,
            'endColumnIndex': len(COURSE_COLS)
        },
        'fields': 'userEnteredValue,userEnteredFormat'
    }
}]})

# 從第2列開始寫入（A2）
ws_course.update(write_rows, 'A2', value_input_option='RAW')

# 標題列格式（第2列，index=1，黑底白字粗體+置中）
sh.batch_update({'requests': [{
    'repeatCell': {
        'range': {
            'sheetId': ws_course.id,
            'startRowIndex': 1,  # 第2列
            'endRowIndex': 2,
            'startColumnIndex': 0,
            'endColumnIndex': len(COURSE_COLS)
        },
        'cell': {
            'userEnteredFormat': {
                'backgroundColor': {'red': 0, 'green': 0, 'blue': 0},
                'textFormat': {
                    'foregroundColor': {'red': 1, 'green': 1, 'blue': 1},
                    'bold': True
                },
                'horizontalAlignment': 'CENTER',
                'verticalAlignment': 'MIDDLE'
            }
        },
        'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)'
    }
}]})

# 凍結第一列（標題）
ws_course.freeze(rows=2)  # 凍結前兩列（篩選按鈕+標題）

# ===== 分隔列填黑色背景 =====
print("設定分隔列顏色...")
separator_rows = []
for i, item in enumerate(all_sheet_rows, 3):  # 從第3列開始（第1列篩選鈕、第2列標題）
    if item['_is_separator']:
        separator_rows.append(i)

if separator_rows:
    # 用 batchUpdate 設定背景色
    sheet_id = ws_course.id
    requests = []
    for row_idx in separator_rows:
        requests.append({
            'repeatCell': {
                'range': {
                    'sheetId': sheet_id,
                    'startRowIndex': row_idx - 1,
                    'endRowIndex': row_idx,
                    'startColumnIndex': 0,
                    'endColumnIndex': len(COURSE_COLS)
                },
                'cell': {
                    'userEnteredFormat': {
                        'backgroundColor': {'red': 0.2, 'green': 0.2, 'blue': 0.2},
                        'textFormat': {
                            'foregroundColor': {'red': 1, 'green': 1, 'blue': 1},
                            'bold': True
                        }
                    }
                },
                'fields': 'userEnteredFormat(backgroundColor,textFormat)'
            }
        })
    sh.batch_update({'requests': requests})
    print(f"OK 分隔列填色完成（{len(separator_rows)} 列）")

# ===== 儲存格填色規則 =====
print("設定儲存格填色...")

# 顏色定義
COLOR_YELLOW_LIGHT = {'red': 1.0,  'green': 0.851, 'blue': 0.4}   # 淡黃（lv1/lv2） #ffd966
COLOR_YELLOW_BRIGHT= {'red': 1.0,  'green': 1.0,   'blue': 0.0}   # 亮黃（專屬/含） #ffff00
COLOR_ORANGE   = {'red': 1.0,  'green': 0.6,  'blue': 0.2}    # 橘色（lv3/lv4）
COLOR_TEAL     = {'red': 0.2,  'green': 0.8,  'blue': 0.8}    # 藍綠色（雙板）
COLOR_WHITE    = {'red': 1.0,  'green': 1.0,  'blue': 1.0}    # 白色
# 整列底色
COLOR_ROW_BLUE   = {'red': 0.8,  'green': 0.9,  'blue': 1.0}  # 淡藍（斑尾）
COLOR_ROW_ORANGE = {'red': 1.0,  'green': 0.9,  'blue': 0.8}  # 淡橘（湯澤）
COLOR_ROW_BROWN  = {'red': 0xcd/255, 'green': 0xa5/255, 'blue': 0x81/255} # 淡棕（龍平）#cda581

# 欄位索引
col_index = {col: i for i, col in enumerate(COURSE_COLS)}
IDX_LEVEL  = col_index.get('LEVEL', -1)
IDX_LEI    = col_index.get('類別', -1)
IDX_TIAN   = col_index.get('天數', -1)
IDX_ZHU    = col_index.get('註記', -1)

sheet_id = ws_course.id
cell_requests = []

def make_row_color(row_idx, color):
    """整列填色"""
    return {
        'repeatCell': {
            'range': {
                'sheetId': sheet_id,
                'startRowIndex': row_idx - 1,
                'endRowIndex': row_idx,
                'startColumnIndex': 0,
                'endColumnIndex': len(COURSE_COLS)
            },
            'cell': {'userEnteredFormat': {'backgroundColor': color}},
            'fields': 'userEnteredFormat.backgroundColor'
        }
    }

def make_cell_color(row_idx, col_idx, color):
    """單格填色"""
    return {
        'repeatCell': {
            'range': {
                'sheetId': sheet_id,
                'startRowIndex': row_idx - 1,
                'endRowIndex': row_idx,
                'startColumnIndex': col_idx,
                'endColumnIndex': col_idx + 1
            },
            'cell': {'userEnteredFormat': {'backgroundColor': color}},
            'fields': 'userEnteredFormat.backgroundColor'
        }
    }

# 逐列檢查資料列（非分隔列、非空列）
data_row_idx = 3  # 從第3列開始（第1列篩選鈕、第2列標題）
for item in all_sheet_rows:
    if item['_is_separator']:
        data_row_idx += 1
        continue

    r = item['_data']

    # 空列（沒有資料的日期）直接跳過，但要計數
    if not r.get('訂編', ''):
        data_row_idx += 1
        continue

    # ── 第1層：整列底色（依項目）──
    item_val = str(r.get('項目', '')).strip()
    if '斑尾' in item_val:
        cell_requests.append(make_row_color(data_row_idx, COLOR_ROW_BLUE))
    elif '湯澤' in item_val:
        cell_requests.append(make_row_color(data_row_idx, COLOR_ROW_ORANGE))
    elif '龍平' in item_val:
        cell_requests.append(make_row_color(data_row_idx, COLOR_ROW_BROWN))
    else:
        cell_requests.append(make_row_color(data_row_idx, COLOR_WHITE))

    # ── 第2層：單格疊加色 ──

    # LEVEL 填色
    lv = str(r.get('LEVEL', '')).strip().lower()
    if IDX_LEVEL >= 0:
        if lv in ('lv1', 'lv2'):
            cell_requests.append(make_cell_color(data_row_idx, IDX_LEVEL, COLOR_YELLOW_LIGHT))
        elif lv in ('lv3', 'lv4'):
            cell_requests.append(make_cell_color(data_row_idx, IDX_LEVEL, COLOR_ORANGE))

    # 類別 = 雙板 → 藍綠色
    lei = str(r.get('類別', '')).strip()
    if IDX_LEI >= 0 and lei == '雙板':
        cell_requests.append(make_cell_color(data_row_idx, IDX_LEI, COLOR_TEAL))

    # 天數含「含」→ 亮黃色
    tian = str(r.get('天數', '')).strip()
    if IDX_TIAN >= 0 and '含' in tian:
        cell_requests.append(make_cell_color(data_row_idx, IDX_TIAN, COLOR_YELLOW_BRIGHT))

    # 註記 = 專屬 → 亮黃色
    zhu = str(r.get('註記', '')).strip()
    if IDX_ZHU >= 0 and zhu == '專屬':
        cell_requests.append(make_cell_color(data_row_idx, IDX_ZHU, COLOR_YELLOW_BRIGHT))

    data_row_idx += 1

# 批次送出（每次最多500個請求）
if cell_requests:
    for i in range(0, len(cell_requests), 500):
        batch = cell_requests[i:i+500]
        sh.batch_update({'requests': batch})
    print(f"OK 儲存格填色完成（{len(cell_requests)} 格）")

# ===== 置中格式 =====
CENTER_COLS = {
    '編號', '出發日期', '曜日', '泊數', '天數', '註記', '訂編', '中文姓名', '性別', '年齡',
    '尺寸', '身高', '體重', '腳長', '雪板', 'LEVEL', '類別',
    '衣褲', '護膝', '護臀', '手套', '專屬', '4件組',
    '教練', '助教', 'checking', '備註'
}

center_requests = []
for col_name, col_idx in col_index.items():
    if col_name in CENTER_COLS:
        center_requests.append({
            'repeatCell': {
                'range': {
                    'sheetId': ws_course.id,
                    'startRowIndex': 1,   # 第2列（標題列）
                    'endRowIndex': len(write_rows) + 1,  # +1因為從第2列起算
                    'startColumnIndex': col_idx,
                    'endColumnIndex': col_idx + 1
                },
                'cell': {
                    'userEnteredFormat': {
                        'horizontalAlignment': 'CENTER',
                        'verticalAlignment': 'MIDDLE'
                    }
                },
                'fields': 'userEnteredFormat(horizontalAlignment,verticalAlignment)'
            }
        })

if center_requests:
    for i in range(0, len(center_requests), 500):
        sh.batch_update({'requests': center_requests[i:i+500]})
    print(f"OK 置中格式設定完成")

# ===== 加黑色格線 =====
print("設定格線...")
total_rows = len(write_rows) + 1  # +1 因為從第2列(index=1)開始寫
total_cols = len(COURSE_COLS)
border_style = {
    'style': 'SOLID',
    'colorStyle': {'rgbColor': {'red': 0, 'green': 0, 'blue': 0}}
}
border_request = {
    'updateBorders': {
        'range': {
            'sheetId': sheet_id,
            'startRowIndex': 1,  # 從第2列開始（不含第1列篩選按鈕）
            'endRowIndex': total_rows,
            'startColumnIndex': 0,
            'endColumnIndex': total_cols
        },
        'top':    border_style,
        'bottom': border_style,
        'left':   border_style,
        'right':  border_style,
        'innerHorizontal': border_style,
        'innerVertical':   border_style,
    }
}
sh.batch_update({'requests': [border_request,
    {
        'updateBorders': {
            'range': {
                'sheetId': sheet_id,
                'startRowIndex': 1,
                'endRowIndex': total_rows,
                'startColumnIndex': total_cols - 1,
                'endColumnIndex': total_cols
            },
            'right': border_style,
        }
    }
]})
print(f"OK 格線設定完成")
print(f"OK 課程安排寫入完成（{len(all_records_flat)} 筆資料，{len(date_map)} 個日期）")

# ===== 重建篩選器（確保範圍涵蓋最新資料）=====
print("重建課程安排篩選器...")
rebuild_filter(ws_course, len(write_rows) - 1, len(COURSE_COLS))

# ===== 隱藏旅客編號和項目1欄位 =====
hidden_col_requests = []
for col_name in HIDDEN_COLS:
    if col_name in COURSE_COLS:
        col_idx = COURSE_COLS.index(col_name)
        hidden_col_requests.append({
            'updateDimensionProperties': {
                'range': {
                    'sheetId': ws_course.id,
                    'dimension': 'COLUMNS',
                    'startIndex': col_idx,
                    'endIndex': col_idx + 1
                },
                'properties': {'hiddenByUser': True},
                'fields': 'hiddenByUser'
            }
        })
if hidden_col_requests:
    sh.batch_update({'requests': hidden_col_requests})
    print(f"OK 隱藏欄位設定完成（{', '.join(HIDDEN_COLS)}）")

# ===== 寫入房務表 =====
def write_sheet(ws_target, cols, headers, manual_cols, important_cols, center_cols, label, exclude_items=None):
    """通用分頁寫入函數（課程安排/房務表共用邏輯）"""
    print(f"\n寫入{label}...")
    exclude_items = exclude_items or set()
    is_room = label == '房務表'

    # 讀取現有資料建立索引（第1列篩選按鈕，第2列標題，第3列起資料）
    existing_data = ws_target.get_all_values()
    existing_map = {}
    if len(existing_data) > 1:
        hrow = existing_data[1]  # 第2列是標題
        cmap = {h: i for i, h in enumerate(hrow)}
        idx_d = cmap.get('訂編', -1)
        idx_n = cmap.get('中文姓名', -1)
        for row in existing_data[2:]:  # 第3列起是資料
            ding = row[idx_d].strip() if idx_d >= 0 and idx_d < len(row) else ''
            name = row[idx_n].strip() if idx_n >= 0 and idx_n < len(row) else ''
            if not ding or not ding.isdigit(): continue
            if not name: continue
            key = f"{ding}_{name}"
            fields = {h: row[i] for h, i in cmap.items() if i < len(row)}
            existing_map[key] = fields
    print(f"  現有：{len(existing_map)} 筆")

    # 建立寫入列（不含標題，標題已在第2列）
    room_write_rows = []
    for item in all_sheet_rows:
        if item['_is_separator']:
            sep = [''] * len(cols)
            sep[1] = item['_data'][1]  # 日期（第二欄）
            sep[2] = item['_data'][2]  # 曜日（第三欄）
            if is_room:
                # 房務表：野:X 斑:X 龍:X（不顯示湯澤）
                counts_raw = item['_data']
                ya = counts_raw[3].replace('野:', '') if len(counts_raw) > 3 and counts_raw[3].startswith('野:') else '0'
                ban = counts_raw[4].replace('斑:', '') if len(counts_raw) > 4 and counts_raw[4].startswith('斑:') else '0'
                lon = counts_raw[6].replace('龍:', '') if len(counts_raw) > 6 and counts_raw[6].startswith('龍:') else '0'
                sep[3] = f"野:{ya}"
                sep[4] = f"斑:{ban}"
                sep[5] = f"龍:{lon}"
            else:
                sep[3] = item['_data'][3]
                sep[4] = item['_data'][4]
                sep[5] = item['_data'][5]
                sep[6] = item['_data'][6]
            room_write_rows.append(sep)
        else:
            r = item['_data']
            if not r.get('訂編',''):
                # 本來就是空列（無資料日期）
                room_write_rows.append([''] * len(cols))
                continue
            # 過濾不寫入的項目（如湯澤純課）
            if r.get('項目','') in exclude_items:
                continue
            key = f"{r.get('訂編','')}_{r.get('中文姓名','')}"
            if key in existing_map:
                old = existing_map[key]
                for mc in manual_cols:
                    if mc in old and old[mc]:
                        r[mc] = old[mc]
            # 房務表：備註初始值來自_房務備註，但不修改原始r
            if is_room:
                room_r = dict(r)
                if not room_r.get('備註','') and room_r.get('_房務備註',''):
                    room_r['備註'] = room_r['_房務備註']
                row_data = [str(room_r.get(c,'')) if room_r.get(c,'') is not None else '' for c in cols]
            else:
                row_data = [str(r.get(c,'')) if r.get(c,'') is not None else '' for c in cols]
            room_write_rows.append(row_data)

    # 補空列：有資料的日期，若過濾後變成只有分隔列（沒有資料列），補一列空列
    final_rows = []
    i = 0
    while i < len(room_write_rows):
        row = room_write_rows[i]
        final_rows.append(row)
        # 如果這列是分隔列
        if row and len(row) > 2 and row[2] in ('一','二','三','四','五','六','日'):
            # 看下一列是否也是分隔列或結尾
            next_i = i + 1
            if next_i >= len(room_write_rows) or (room_write_rows[next_i] and len(room_write_rows[next_i]) > 2 and room_write_rows[next_i][2] in ('一','二','三','四','五','六','日')):
                # 補空列
                final_rows.append([''] * len(cols))
        i += 1
    room_write_rows = final_rows

    # 清除第3列以後的內容和格式（保留第1列篩選按鈕、第2列標題不動）
    sh.batch_update({'requests': [{
        'updateCells': {
            'range': {'sheetId': ws_target.id, 'startRowIndex': 2, 'endRowIndex': 2000,
                      'startColumnIndex': 0, 'endColumnIndex': len(cols)},
            'fields': 'userEnteredValue,userEnteredFormat'
        }
    }]})

    # 從第3列開始寫入資料（跳過標題列，標題已在第2列）
    ws_target.update(room_write_rows, 'A3', value_input_option='RAW')

    sid = ws_target.id
    reqs = []

    # 標題列黑底白字置中（第2列，index=1）
    reqs.append({'repeatCell': {
        'range': {'sheetId': sid, 'startRowIndex': 1, 'endRowIndex': 2, 'startColumnIndex': 0, 'endColumnIndex': len(cols)},
        'cell': {'userEnteredFormat': {
            'backgroundColor': {'red':0,'green':0,'blue':0},
            'textFormat': {'foregroundColor': {'red':1,'green':1,'blue':1}, 'bold': True},
            'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE'
        }},
        'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)'
    }})

    # 分隔列黑底（用room_write_rows直接計算位置）
    sep_rows = []
    for i, row in enumerate(room_write_rows, 3):  # 從第3列開始（跳過篩選鈕+標題）
        if row and len(row) > 2 and row[2] in ('一','二','三','四','五','六','日'):
            sep_rows.append(i)
    for ri in sep_rows:
        reqs.append({'repeatCell': {
            'range': {'sheetId': sid, 'startRowIndex': ri-1, 'endRowIndex': ri, 'startColumnIndex': 0, 'endColumnIndex': len(cols)},
            'cell': {'userEnteredFormat': {
                'backgroundColor': {'red':0.2,'green':0.2,'blue':0.2},
                'textFormat': {'foregroundColor': {'red':1,'green':1,'blue':1}, 'bold': True}
            }},
            'fields': 'userEnteredFormat(backgroundColor,textFormat)'
        }})

    # 整列底色
    col_idx_map = {c: i for i, c in enumerate(cols)}
    idx_pou = col_idx_map.get('泊數', -1)
    idx_item = col_idx_map.get('項目', -1)

    for ridx, row in enumerate(room_write_rows, 3):  # 從第3列開始
        if not row: continue
        if len(row) > 2 and row[2] in ('一','二','三','四','五','六','日'): continue
        if all(v == '' for v in row): continue

        item_val = row[idx_item].strip() if idx_item >= 0 and idx_item < len(row) else ''
        if not item_val: continue

        if '斑尾' in item_val:    bg = COLOR_ROW_BLUE
        elif '湯澤' in item_val:  bg = COLOR_ROW_ORANGE
        elif '龍平' in item_val:  bg = COLOR_ROW_BROWN
        else:                     bg = COLOR_WHITE

        reqs.append({'repeatCell': {
            'range': {'sheetId': sid, 'startRowIndex': ridx-1, 'endRowIndex': ridx, 'startColumnIndex': 0, 'endColumnIndex': len(cols)},
            'cell': {'userEnteredFormat': {'backgroundColor': bg}},
            'fields': 'userEnteredFormat.backgroundColor'
        }})

        # 泊數不是3的填黃色
        if is_room and idx_pou >= 0 and idx_pou < len(row):
            pou_val = row[idx_pou].strip()
            if pou_val and pou_val != '3':
                reqs.append({'repeatCell': {
                    'range': {'sheetId': sid, 'startRowIndex': ridx-1, 'endRowIndex': ridx, 'startColumnIndex': idx_pou, 'endColumnIndex': idx_pou+1},
                    'cell': {'userEnteredFormat': {'backgroundColor': COLOR_YELLOW_BRIGHT}},
                    'fields': 'userEnteredFormat.backgroundColor'
                }})

    # 格線（從第2列標題開始）
    reqs.append({'updateBorders': {
        'range': {'sheetId': sid, 'startRowIndex': 1, 'endRowIndex': len(room_write_rows)+2, 'startColumnIndex': 0, 'endColumnIndex': len(cols)},
        'top': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'bottom': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'left': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'right': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'innerHorizontal': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'innerVertical': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
    }})
    # 確保最右欄右側格線
    reqs.append({'updateBorders': {
        'range': {'sheetId': sid, 'startRowIndex': 1, 'endRowIndex': len(room_write_rows)+2, 'startColumnIndex': len(cols)-1, 'endColumnIndex': len(cols)},
        'right': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
    }})

    # 置中（從第2列標題開始）
    for cn, ci in col_idx_map.items():
        if cn in center_cols:
            reqs.append({'repeatCell': {
                'range': {'sheetId': sid, 'startRowIndex': 1, 'endRowIndex': len(room_write_rows)+2, 'startColumnIndex': ci, 'endColumnIndex': ci+1},
                'cell': {'userEnteredFormat': {'horizontalAlignment':'CENTER','verticalAlignment':'MIDDLE'}},
                'fields': 'userEnteredFormat(horizontalAlignment,verticalAlignment)'
            }})

    # 凍結前兩列（篩選按鈕+標題）
    ws_target.freeze(rows=2)

    # 自動欄寬（依欄位名稱估算）
    COL_WIDTH = {
        # 課程安排
        '編號': 30, '泊數': 30, '天數': 30, '性別': 30, '年齡': 30,
        '尺寸': 30, '身高': 30, '體重': 30, '腳長': 30, '雪板': 30,
        'Lv': 30, 'LEVEL': 30, '類別': 30,
        '衣褲': 30, '護膝': 30, '護臀': 30, '手套': 30, '四件': 30, '4件組': 30,
        '註記': 35, '訂編': 35,
        '項目': 60, '出發日期': 60, '中文姓名': 60,
        '英文姓名': 150,
        '教練': 70, '助教': 70, 'checking': 70,
        '備註': 70,
        # 房務表
        '房號': 40, '序號': 40,
        '飲食': 80,
        'OP備註': 120,
        # 共用（房務表用不同寬度時覆蓋）
    }
    # 房務表特殊欄寬
    ROOM_COL_WIDTH = {
        '項目': 40, '泊數': 40, '訂編': 40, '性別': 40, '年齡': 40,
        '房號': 40, '序號': 40,
        '出發日期': 70, '中文姓名': 70,
        '英文姓名': 130,
        '備註': 130,
        '飲食': 80,
        'OP備註': 120,
    }
    effective_width = ROOM_COL_WIDTH if is_room else COL_WIDTH
    col_width_reqs = []
    for ci, cn in enumerate(cols):
        px = effective_width.get(cn, COL_WIDTH.get(cn, 80))
        col_width_reqs.append({
            'updateDimensionProperties': {
                'range': {'sheetId': sid, 'dimension': 'COLUMNS', 'startIndex': ci, 'endIndex': ci+1},
                'properties': {'pixelSize': px},
                'fields': 'pixelSize'
            }
        })
    # 列高：標題列30，其他24
    col_width_reqs.append({
        'updateDimensionProperties': {
            'range': {'sheetId': sid, 'dimension': 'ROWS', 'startIndex': 0, 'endIndex': 1},
            'properties': {'pixelSize': 30},
            'fields': 'pixelSize'
        }
    })
    col_width_reqs.append({
        'updateDimensionProperties': {
            'range': {'sheetId': sid, 'dimension': 'ROWS', 'startIndex': 1, 'endIndex': len(room_write_rows)},
            'properties': {'pixelSize': 24},
            'fields': 'pixelSize'
        }
    })
    reqs.extend(col_width_reqs)

    # 批次送出
    for i in range(0, len(reqs), 500):
        sh.batch_update({'requests': reqs[i:i+500]})

    print(f"  OK {label}寫入完成（{len(all_records_flat)} 筆）")

    # 重建篩選器
    print(f"  重建{label}篩選器...")
    rebuild_filter(ws_target, len(room_write_rows), len(cols))

# 確保房務表分頁存在
try:
    ws_room = sh.worksheet(SHEET_ROOM)
except gspread.exceptions.WorksheetNotFound:
    ws_room = sh.add_worksheet(title=SHEET_ROOM, rows=2000, cols=25)
    print(f"建立新分頁：{SHEET_ROOM}")

write_sheet(ws_room, ROOM_COLS, ROOM_HEADERS, ROOM_MANUAL_COLS, ROOM_IMPORTANT_COLS, ROOM_CENTER_COLS, '房務表', exclude_items=ROOM_EXCLUDE_ITEMS)

# ===== 寫入三個入住單 =====
for ci_sheet in CHECKIN_SHEETS:
    # 用gid找分頁，避免名稱改變導致重複建立
    ws_ci = None
    for ws in sh.worksheets():
        if ws.id == ci_sheet['gid']:
            ws_ci = ws
            break
    if ws_ci is None:
        ws_ci = sh.add_worksheet(title=ci_sheet['name'], rows=2000, cols=20)
        print(f"建立新分頁：{ci_sheet['name']}")
    else:
        print(f"找到分頁：{ws_ci.title}（gid={ws_ci.id}）")

    # 用 include 函數過濾，只取該區資料
    include_fn = ci_sheet['include']

    # 讀取現有資料（第1列是標題，第2列起是資料）
    ci_existing = ws_ci.get_all_values()
    ci_map = {}
    if len(ci_existing) > 0:
        ci_hrow = ci_existing[0]  # 第1列是標題
        ci_cmap = {h: i for i, h in enumerate(ci_hrow)}
        ci_visitor_idx = ci_cmap.get('旅客編號', -1)
        ci_ding_idx = ci_cmap.get('訂編', -1)
        ci_name_idx = ci_cmap.get('中文姓名', -1)
        for row in ci_existing[1:]:  # 第2列起是資料
            if not row: continue
            visitor_id = row[ci_visitor_idx].strip() if ci_visitor_idx >= 0 and ci_visitor_idx < len(row) else ''
            ding = row[ci_ding_idx].strip() if ci_ding_idx >= 0 and ci_ding_idx < len(row) else ''
            name = row[ci_name_idx].strip() if ci_name_idx >= 0 and ci_name_idx < len(row) else ''
            if visitor_id:
                key = f"v_{visitor_id}"
            elif ding and name:
                key = f"{ding}_{name}"
            else:
                continue
            ci_map[key] = {h: row[i] for h, i in ci_cmap.items() if i < len(row)}

    # 建立寫入列
    ci_write_rows = []
    for item in all_sheet_rows:
        if item['_is_separator']:
            # 入住單分隔列：只留日期和曜日
            sep = [''] * len(CHECKIN_COLS)
            sep[0] = item['_data'][1]  # 日期（第一欄）
            sep[1] = item['_data'][2]  # 曜日（第二欄）
            ci_write_rows.append(sep)
        else:
            r = item['_data']
            if not r.get('訂編', ''): continue
            item_val = r.get('項目', '')
            if not include_fn(item_val): continue

            # 取key
            visitor_id = r.get('旅客編號', '').strip()
            key = f"v_{visitor_id}" if visitor_id else f"{r.get('訂編','')}_{r.get('中文姓名','')}"

            # 保留手動欄位
            if key in ci_map:
                old = ci_map[key]
                for mc in CHECKIN_MANUAL_COLS:
                    if mc in old and old[mc]:
                        r[mc] = old[mc]

            row_data = []
            for col in CHECKIN_COLS:
                v = r.get(col, '')
                row_data.append(str(v) if v is not None else '')
            ci_write_rows.append(row_data)

    # 如果完全沒有資料列，補一個空的分隔列佔位
    if not ci_write_rows:
        ci_write_rows.append([''] * len(CHECKIN_COLS))

    # 清除第2列以後，保留第1列標題
    sh.batch_update({'requests': [{
        'updateCells': {
            'range': {
                'sheetId': ws_ci.id,
                'startRowIndex': 1, 'endRowIndex': 2000,
                'startColumnIndex': 0, 'endColumnIndex': len(CHECKIN_COLS)
            },
            'fields': 'userEnteredValue,userEnteredFormat'
        }
    }]})
    # 寫入標題列（第1列）
    ws_ci.update([CHECKIN_HEADERS], 'A1', value_input_option='RAW')
    # 寫入資料（從第2列）
    if ci_write_rows:
        ws_ci.update(ci_write_rows, 'A2', value_input_option='RAW')

    # 格式設定
    ci_sid = ws_ci.id
    ci_reqs = []

    # 標題列（第1列，index=0）黑底白字
    ci_reqs.append({'repeatCell': {
        'range': {'sheetId': ci_sid, 'startRowIndex': 0, 'endRowIndex': 1,
                  'startColumnIndex': 0, 'endColumnIndex': len(CHECKIN_COLS)},
        'cell': {'userEnteredFormat': {
            'backgroundColor': {'red': 0, 'green': 0, 'blue': 0},
            'textFormat': {'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}, 'bold': True},
            'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE'
        }},
        'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)'
    }})

    # 整列底色 + 分隔列填色（從第2列，index=1起）
    col_map_ci = {c: i for i, c in enumerate(CHECKIN_COLS)}
    ridx = 2  # 資料從第2列開始
    for item in all_sheet_rows:
        if item['_is_separator']:
            ci_reqs.append({'repeatCell': {
                'range': {'sheetId': ci_sid, 'startRowIndex': ridx-1, 'endRowIndex': ridx,
                          'startColumnIndex': 0, 'endColumnIndex': len(CHECKIN_COLS)},
                'cell': {'userEnteredFormat': {
                    'backgroundColor': {'red': 0.2, 'green': 0.2, 'blue': 0.2},
                    'textFormat': {'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}, 'bold': True}
                }},
                'fields': 'userEnteredFormat(backgroundColor,textFormat)'
            }})
            ridx += 1
        else:
            r = item['_data']
            if not r.get('訂編', ''): continue
            item_val = r.get('項目', '')
            if not include_fn(item_val): continue

            if '斑尾' in item_val:    bg = COLOR_ROW_BLUE
            elif '湯澤' in item_val:  bg = COLOR_ROW_ORANGE
            elif '龍平' in item_val:  bg = COLOR_ROW_BROWN
            else:                     bg = COLOR_WHITE
            ci_reqs.append({'repeatCell': {
                'range': {'sheetId': ci_sid, 'startRowIndex': ridx-1, 'endRowIndex': ridx,
                          'startColumnIndex': 0, 'endColumnIndex': len(CHECKIN_COLS)},
                'cell': {'userEnteredFormat': {'backgroundColor': bg}},
                'fields': 'userEnteredFormat.backgroundColor'
            }})
            ridx += 1

    # 格線
    total_ci_rows = ridx - 1
    ci_reqs.append({'updateBorders': {
        'range': {'sheetId': ci_sid, 'startRowIndex': 0, 'endRowIndex': total_ci_rows,
                  'startColumnIndex': 0, 'endColumnIndex': len(CHECKIN_COLS)},
        'top': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'bottom': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'left': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'right': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'innerHorizontal': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
        'innerVertical': {'style':'SOLID','colorStyle':{'rgbColor':{'red':0,'green':0,'blue':0}}},
    }})

    # 置中
    for cn, ci in col_map_ci.items():
        if cn in CHECKIN_CENTER_COLS:
            ci_reqs.append({'repeatCell': {
                'range': {'sheetId': ci_sid, 'startRowIndex': 0, 'endRowIndex': total_ci_rows,
                          'startColumnIndex': ci, 'endColumnIndex': ci+1},
                'cell': {'userEnteredFormat': {'horizontalAlignment':'CENTER','verticalAlignment':'MIDDLE'}},
                'fields': 'userEnteredFormat(horizontalAlignment,verticalAlignment)'
            }})

    # 凍結第一列
    ws_ci.freeze(rows=1)

    # 欄寬
    CI_COL_WIDTHS = {
        '項目': 70, '出發日期': 65, '泊數': 40, '天數': 50,
        '訂編': 45, '中文姓名': 70, '英文姓名': 130,
        '性別': 35, '年齡': 35, '房型': 80, '入住備註': 150
    }
    for ci_idx, cn in enumerate(CHECKIN_COLS):
        px = CI_COL_WIDTHS.get(cn, 70)
        ci_reqs.append({'updateDimensionProperties': {
            'range': {'sheetId': ci_sid, 'dimension': 'COLUMNS', 'startIndex': ci_idx, 'endIndex': ci_idx+1},
            'properties': {'pixelSize': px}, 'fields': 'pixelSize'
        }})

    # 送出
    for i in range(0, len(ci_reqs), 500):
        sh.batch_update({'requests': ci_reqs[i:i+500]})

    # 篩選器（從第1列標題開始）
    rebuild_filter(ws_ci, len(ci_write_rows), len(CHECKIN_COLS), start_row=0)

    print(f"OK {ws_ci.title} 寫入完成")

# ===== 完整大表.xlsx（本機留存）=====
FULL_COLS = [
    '項目', 'OP備註', '出發日期', '泊數', '天數', '註記', '訂編',
    '中文姓名', '英文姓名', '性別', '年齡',
    '尺寸', '身高', '體重', '腳長', '雪板',
    'LEVEL', '類別', '衣褲', '護膝', '護臀', '手套', '4件組',
    '教練', '助教', 'checking', '備註', '報名日期', '團費',
    '飲食', '雪場房號', '序號', '雪場備註', '市區房號', '序號.1', '市區備註',
    '手機號碼', 'EMail', '國籍', '身分證', '生日',
    '護照號碼', '護照到期日', '旅客編號',
]

wb = Workbook()
ws = wb.active
ws.title = '完整大表'
header_font = Font(bold=True)
for col_idx, col_name in enumerate(FULL_COLS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font

BLACK_FILL = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
row_num = 2
prev_date = None

for r in all_records_flat:
    cur_date = r.get('出發日期','')
    if cur_date != prev_date:
        for col_idx in range(1, len(FULL_COLS)+1):
            ws.cell(row=row_num, column=col_idx, value='').fill = BLACK_FILL
        row_num += 1
        prev_date = cur_date
    for col_idx, col_name in enumerate(FULL_COLS, 1):
        v = clean_val(r.get(col_name,''))
        ws.cell(row=row_num, column=col_idx, value=v)
    row_num += 1

for i in range(1, len(FULL_COLS)+1):
    ws.column_dimensions[get_column_letter(i)].width = 12

full_path = os.path.join(folder, '完整大表.xlsx')
wb.save(full_path)
print(f"OK 完整大表.xlsx 輸出完成（本機留存，含個資）")

# ===== 每日教練需求報表 =====
print("\n生成每日教練需求報表...")
try:
    import json as _json, re as _re2, math as _math
    from collections import defaultdict as _dd
    from datetime import timedelta as _td

    PURE_ITEMS_C = {'野澤純課','斑尾純課','湯澤純課'}
    AREA_MAP_C = {'野澤純課':'野澤','斑尾純課':'斑尾','湯澤純課':'湯澤',
                  '野澤':'野澤','斑尾':'斑尾','龍平':'龍平'}

    def _lv_grp(lv):
        if not lv or lv == 'lv0': return 'g0'
        n = int(lv.replace('lv','')) if lv.startswith('lv') else 0
        if n <= 2: return 'g12'
        return 'g345'

    def _parse_days(tiansu):
        """從天數欄解析實際上課天數（含教練的部分）"""
        if not tiansu or str(tiansu).strip() == '': return 2
        s = str(tiansu).strip()
        # 取數字部分
        m = _re2.search(r'(\d+)', s)
        if not m: return 2
        n = int(m.group(1))
        # 「含」代表有含教練的續滑，天數已包含
        # 「無」代表不含教練的續滑，教練只上基本天數
        if '無' in s and '含' not in s:
            # 純不含教練續滑，教練只需要基本2天
            return 2
        return n  # 含教練或純課程，直接用天數

    # 每日計算（從all_records_flat）
    _daily = _dd(lambda: _dd(lambda: {'pure_orders':set(),'g0':0,'g12':0,'g345':0}))

    for r in all_records_flat:
        item = r.get('項目','')
        if not item or item == '其他': continue
        area = AREA_MAP_C.get(item,'')
        if not area: continue

        dep_str = r.get('出發日期','')
        if not dep_str: continue
        try:
            parts = dep_str.split('/')
            if len(parts) == 3:
                dep = datetime(2000+int(parts[0]), int(parts[1]), int(parts[2]))
            else:
                continue
        except:
            continue

        ono = str(r.get('訂編','')).strip()
        lv = r.get('LEVEL','') or 'lv0'
        grp = _lv_grp(lv)
        tiansu = str(r.get('天數','')).strip()

        if item in PURE_ITEMS_C:
            # 純課程：從天數欄取天數
            days = _parse_days(tiansu)
            if days == 0: days = 2
            for d in range(1, days+1):
                ds = (dep + _td(days=d)).strftime('%Y/%m/%d')
                _daily[ds][area]['pure_orders'].add(ono)
        else:
            # 團客：從天數欄判斷含教練天數
            days = _parse_days(tiansu)
            if days == 0: days = 2
            for d in range(1, days+1):
                ds = (dep + _td(days=d)).strftime('%Y/%m/%d')
                _daily[ds][area][grp] += 1

    def _calc(d):
        g0,g12,g345 = d.get('g0',0),d.get('g12',0),d.get('g345',0)
        pure = len(d.get('pure_orders',set()))
        team = (_math.ceil(g0/6) if g0>0 else 0) + (_math.ceil(g12/6) if g12>0 else 0) + (_math.ceil(g345/6) if g345>0 else 0)
        return pure+team, pure, team, g0, g12, g345

    # 整理輸出資料
    AREAS_C = ['野澤','斑尾','湯澤','龍平']
    CAPS = {'野澤':14,'斑尾':8,'湯澤':10,'龍平':10}
    THRESHOLDS = {
        '野澤': [(6,'green'),(9,'yellow'),(11,'orange'),(999,'red')],
        '斑尾': [(4,'green'),(6,'yellow'),(7,'orange'),(999,'red')],
        '湯澤': [(5,'green'),(7,'yellow'),(9,'orange'),(999,'red')],
        '龍平': [(5,'green'),(7,'yellow'),(9,'orange'),(999,'red')],
    }

    def _color(area, n):
        if n == 0: return 'none'
        for limit, color in THRESHOLDS[area]:
            if n <= limit: return color
        return 'red'

    # 雪季日期範圍
    _start = datetime(DATE_START.year, DATE_START.month, DATE_START.day)
    _end = datetime(DATE_END.year, DATE_END.month, DATE_END.day)

    coach_rows = []
    cur = _start
    while cur <= _end:
        ds = cur.strftime('%Y/%m/%d')
        wd = ['一','二','三','四','五','六','日'][cur.weekday()]
        row_data = {'date': ds, 'wd': wd, 'areas': {}}
        total_jp, total_kr, total_all = 0, 0, 0
        for area in AREAS_C:
            d = _daily[ds].get(area, {})
            coaches, pure, team, g0, g12, g345 = _calc(d)
            color = _color(area, coaches)
            cap = CAPS[area]
            row_data['areas'][area] = {
                'c': coaches, 'pure': pure, 'team': team,
                'g0': g0, 'g12': g12, 'g345': g345,
                'color': color, 'cap': cap
            }
            if area in ['野澤','斑尾','湯澤']: total_jp += coaches
            else: total_kr += coaches
            total_all += coaches
        row_data['total_jp'] = total_jp
        row_data['total_kr'] = total_kr
        row_data['total'] = total_all
        coach_rows.append(row_data)
        cur += _td(days=1)

    coach_json = _json.dumps(coach_rows, ensure_ascii=False)

    coach_html = r"""<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>DBC 每日教練需求</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@300;400;500;700&family=DM+Mono:wght@400;500&display=swap');
:root{--bg:#0a0f1e;--s1:#111827;--s2:#1a2235;--bd:#1e2d45;--tx:#e2e8f0;--mu:#64748b;
  --ny:#ffffff;--nb:#7ec8f5;--nt:#ffb347;--nl:#c8864a;
  --green:#10b981;--yellow:#f59e0b;--orange:#f97316;--red:#ef4444;--none:#334155;}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--tx);font-family:'Noto Sans TC',sans-serif;font-size:15px}
.hd{background:linear-gradient(135deg,#0f172a,#1e3a5f 50%,#0f172a);border-bottom:1px solid var(--bd);padding:14px 20px;position:sticky;top:0;z-index:200}
.hd-in{max-width:1400px;margin:0 auto;display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap}
.logo{display:flex;align-items:center;gap:10px}
.logo-i{width:30px;height:30px;background:#3b82f6;border-radius:7px;display:flex;align-items:center;justify-content:center}
.logo-t h1{font-size:15px;font-weight:700}
.logo-t p{font-size:10px;color:var(--mu)}
.badge{background:var(--s2);border:1px solid var(--bd);border-radius:20px;padding:4px 12px;font-size:11px;color:#06b6d4;font-family:'DM Mono',monospace}
.main{max-width:1400px;margin:0 auto;padding:16px}
.view-toggle{display:flex;gap:6px;margin-bottom:14px}
.vt-btn{padding:7px 20px;border-radius:7px;font-size:13px;cursor:pointer;border:1px solid var(--bd);color:var(--mu);background:transparent;transition:all .15s;font-family:'Noto Sans TC',sans-serif}
.vt-btn.active{background:#3b82f6;border-color:#3b82f6;color:#fff;font-weight:500}
.legend{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:14px;padding:12px 16px;background:var(--s1);border:1px solid var(--bd);border-radius:8px;align-items:center}
/* 日曆 */
.cal-wrap{display:flex;flex-direction:column;gap:24px}
.cal-month{background:var(--s1);border:1px solid var(--bd);border-radius:10px;overflow:hidden}
.cal-month-title{padding:12px 16px;font-size:15px;font-weight:700;background:var(--s2);border-bottom:1px solid var(--bd);color:#3b82f6}
.cal-grid{display:grid;grid-template-columns:repeat(7,1fr);border-left:1px solid var(--bd)}
.cal-dow{padding:8px 4px;text-align:center;font-size:11px;color:var(--mu);border-bottom:1px solid var(--bd);border-right:1px solid var(--bd);background:var(--s2)}
.cal-day{border-right:1px solid var(--bd);border-bottom:1px solid var(--bd);min-height:90px;padding:6px;position:relative}
.cal-day.empty{background:rgba(0,0,0,.15)}
.cal-day.weekend{background:rgba(245,158,11,.03)}
.cal-day.has-data{background:rgba(30,45,69,.4)}
.cal-date{font-size:12px;color:var(--mu);margin-bottom:6px;font-family:'DM Mono',monospace}
.cal-date.today{color:#3b82f6;font-weight:700}
.cal-badges{display:flex;flex-direction:column;gap:3px}
.cal-badge{display:flex;align-items:center;gap:4px;font-size:11px;border-radius:4px;padding:2px 5px}
.cal-badge-dot{width:6px;height:6px;border-radius:50%;flex-shrink:0}
.cal-badge-n{font-family:'DM Mono',monospace;font-weight:700;font-size:12px}
.cal-badge.bg-green{background:rgba(16,185,129,.1)}
.cal-badge.bg-yellow{background:rgba(245,158,11,.1)}
.cal-badge.bg-orange{background:rgba(249,115,22,.1)}
.cal-badge.bg-red{background:rgba(239,68,68,.1)}
@media(max-width:600px){
  .cal-day{min-height:70px;padding:4px}
  .cal-badge{font-size:10px}
  .cal-badge-n{font-size:11px}
}
.legend-item{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--mu)}
.leg-dot{width:10px;height:10px;border-radius:50%}
.filters{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:14px}
.filter-btn{padding:6px 16px;border-radius:6px;font-size:12px;cursor:pointer;border:1px solid var(--bd);color:var(--mu);background:transparent;transition:all .15s}
.filter-btn.active{border-color:#3b82f6;color:#3b82f6;background:rgba(59,130,246,.1)}
.tbl-wrap{border-radius:10px;border:1px solid var(--bd);overflow:auto;max-height:calc(100vh - 180px)}
table{width:100%;border-collapse:collapse;min-width:600px}
thead{position:sticky;top:0;z-index:100}
thead th{background:#0d1829;padding:12px 16px;text-align:center;font-size:13px;color:var(--mu);font-weight:500;border-bottom:1px solid var(--bd);white-space:nowrap}
thead tr:first-child th{border-bottom:none}
thead tr:last-child th{border-bottom:2px solid var(--bd);font-size:12px}
tbody tr{border-bottom:1px solid var(--bd);transition:background .1s}
tbody tr:hover{background:var(--s2)}
tbody tr.has-data{background:rgba(30,45,69,.3)}
td{padding:12px 16px;text-align:center;font-family:'DM Mono',monospace;font-size:17px}
td.date-col{text-align:left;font-family:'Noto Sans TC',sans-serif;color:var(--mu);font-size:14px;white-space:nowrap}
td.wd-col{font-size:14px;color:var(--mu);padding:12px 6px}
.coach-badge{display:inline-flex;align-items:center;justify-content:center;min-width:48px;height:40px;border-radius:8px;font-weight:700;font-size:22px;padding:0 12px}
.cb-green{background:rgba(16,185,129,.15);color:var(--green)}
.cb-yellow{background:rgba(245,158,11,.15);color:var(--yellow)}
.cb-orange{background:rgba(249,115,22,.15);color:var(--orange)}
.cb-red{background:rgba(239,68,68,.15);color:var(--red)}
.cb-none{color:var(--mu);opacity:.3;font-size:16px}
td.total-jp{font-size:20px;font-weight:700}
td.total-kr{font-size:20px;font-weight:700}
td.total-all{font-size:24px;font-weight:700}
.month-header td{background:rgba(15,23,42,.95);color:#3b82f6;font-size:13px;font-weight:600;text-align:left;padding:8px 16px;border-bottom:1px solid #1e3a5f}
@media(max-width:600px){
  td,thead th{padding:8px 8px;font-size:14px}
  .coach-badge{min-width:38px;height:34px;font-size:18px}
  td.total-all{font-size:20px}
}
</style></head><body>
<div class="hd"><div class="hd-in">
  <div class="logo"><div class="logo-i">👨‍🏫</div><div class="logo-t"><h1>每日教練需求</h1><p>DAILY COACH REQUIREMENT</p></div></div>
  <div class="badge" id="upd">—</div>
</div></div>
<div class="main">

<div class="view-toggle">
  <button class="vt-btn active" onclick="setView('table',this)">📋 列表</button>
  <button class="vt-btn" onclick="setView('calendar',this)">📅 日曆</button>
</div>

<div id="table-view">
<div class="legend">
  <span style="font-size:11px;color:var(--mu);margin-right:4px">教練人數警示：</span>
  <div class="legend-item"><div class="leg-dot" style="background:var(--green)"></div>充足</div>
  <div class="legend-item"><div class="leg-dot" style="background:var(--yellow)"></div>注意</div>
  <div class="legend-item"><div class="leg-dot" style="background:var(--orange)"></div>緊張</div>
  <div class="legend-item"><div class="leg-dot" style="background:var(--red)"></div>超載</div>
  <div style="margin-left:auto;font-size:10px;color:var(--mu)">上限：野澤14 斑尾8 湯澤10 龍平10</div>
</div>

<div class="filters">
  <button class="filter-btn active" onclick="setFilter('all',this)">全部</button>
  <button class="filter-btn" onclick="setFilter('has',this)">有課日</button>
  <button class="filter-btn" onclick="setFilter('warn',this)">注意/緊張/超載</button>
  <button class="filter-btn" onclick="setFilter('nozawa',this)">野澤</button>
  <button class="filter-btn" onclick="setFilter('madarao',this)">斑尾</button>
  <button class="filter-btn" onclick="setFilter('yuzawa',this)">湯澤</button>
  <button class="filter-btn" onclick="setFilter('ryongpyong',this)">龍平</button>
</div>

<div class="tbl-wrap">
<table id="coach-table">
<thead>
  <tr>
    <th rowspan="2" style="text-align:left">日期</th>
    <th rowspan="2" style="padding:10px 4px">曜</th>
    <th style="--area-c:var(--ny);border-bottom:2px solid var(--ny)">野澤</th>
    <th style="--area-c:var(--nb);border-bottom:2px solid var(--nb)">斑尾</th>
    <th style="--area-c:var(--nt);border-bottom:2px solid var(--nt)">湯澤</th>
    <th style="--area-c:var(--nl);border-bottom:2px solid var(--nl)">龍平</th>
    <th rowspan="2">日本<br><span style="font-size:9px;color:var(--mu)">(野+斑+湯)</span></th>
    <th rowspan="2">韓國<br><span style="font-size:9px;color:var(--mu)">(龍平)</span></th>
    <th rowspan="2">合計</th>
  </tr>
  <tr>
    <th></th>
    <th></th>
    <th></th>
    <th></th>
  </tr>
</thead>
<tbody id="tbody"></tbody>
</table>
</div>
</div>
</div><!-- end table-view -->

<div id="cal-view" style="display:none">
  <div class="legend">
    <span style="font-size:11px;color:var(--mu);margin-right:4px">教練人數警示：</span>
    <div class="legend-item"><div class="leg-dot" style="background:var(--green)"></div>充足</div>
    <div class="legend-item"><div class="leg-dot" style="background:var(--yellow)"></div>注意</div>
    <div class="legend-item"><div class="leg-dot" style="background:var(--orange)"></div>緊張</div>
    <div class="legend-item"><div class="leg-dot" style="background:var(--red)"></div>超載</div>
  </div>
  <div class="cal-wrap" id="cal-wrap"></div>
</div>

<script>
const D=COACH_DATA_PLACEHOLDER;
const AREAS=['野澤','斑尾','湯澤','龍平'];
const AC={'野澤':'ny','斑尾':'nb','湯澤':'nt','龍平':'nl'};

document.getElementById('upd').textContent='更新：'+new Date().toLocaleDateString('zh-TW');

function badge(n, color){
  if(n===0) return `<div class="coach-badge cb-none">—</div>`;
  return `<div class="coach-badge cb-${color}">${n}</div>`;
}

let currentFilter='all';
function setFilter(f,btn){
  currentFilter=f;
  document.querySelectorAll('.filter-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  renderTable();
}

function shouldShow(row){
  if(currentFilter==='all') return true;
  if(currentFilter==='has') return row.total>0;
  if(currentFilter==='warn'){
    return AREAS.some(a=>{
      const c=row.areas[a]?.color;
      return c==='yellow'||c==='orange'||c==='red';
    });
  }
  const aMap={'nozawa':'野澤','madarao':'斑尾','yuzawa':'湯澤','ryongpyong':'龍平'};
  const area=aMap[currentFilter];
  return area && (row.areas[area]?.c||0)>0;
}

function renderTable(){
  const tbody=document.getElementById('tbody');
  tbody.innerHTML='';
  let prevMonth='';
  D.forEach(row=>{
    if(!shouldShow(row)) return;
    const month=row.date.slice(0,7);
    if(month!==prevMonth){
      const mtr=document.createElement('tr');
      mtr.className='month-header';
      mtr.innerHTML=`<td colspan="9">${month.replace('2026-','26/').replace('2027-','27/')} ${['一','二','三','四','五','六','日'][new Date(row.date).getDay()==0?6:new Date(row.date).getDay()-1]}</td>`;
      tbody.appendChild(mtr);
      prevMonth=month;
    }
    const isWeekend=row.wd==='六'||row.wd==='日';
    const tr=document.createElement('tr');
    tr.className=(row.total>0?'has-data ':'')+(isWeekend?'weekend':'');
    let cols='';
    AREAS.forEach(a=>{
      const d=row.areas[a]||{c:0,color:'none'};
      cols+=`<td>${badge(d.c,d.color)}</td>`;
    });
    const jpColor=row.total_jp>25?'red':row.total_jp>20?'orange':row.total_jp>15?'yellow':'';
    tr.innerHTML=`
      <td class="date-col">${row.date.slice(5)}</td>
      <td class="wd-col" style="color:${row.wd==='六'||row.wd==='日'?'#f59e0b':'var(--mu)'}">${row.wd}</td>
      ${cols}
      <td class="total-jp" style="color:${jpColor?'var(--'+jpColor+')':'#06b6d4'};font-size:16px;font-weight:700">${row.total_jp||'—'}</td>
      <td class="total-kr" style="color:${row.total_kr>0?'#8b5cf6':'var(--mu)'};font-size:16px;font-weight:700">${row.total_kr||'—'}</td>
      <td class="total-all" style="font-size:18px;font-weight:700">${row.total||'—'}</td>
    `;
    tbody.appendChild(tr);
  });
}
renderTable();

// ── View 切換 ──
function setView(v, btn){
  document.querySelectorAll('.vt-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  document.getElementById('table-view').style.display = v==='table'?'':'none';
  document.getElementById('cal-view').style.display = v==='calendar'?'':'none';
  if(v==='calendar' && !calRendered) renderCalendar();
}

// ── 日曆渲染 ──
let calRendered=false;
const DOW=['日','一','二','三','四','五','六'];
const COLOR_MAP={'green':'var(--green)','yellow':'var(--yellow)','orange':'var(--orange)','red':'var(--red)','none':'var(--mu)'};

function renderCalendar(){
  calRendered=true;
  const wrap=document.getElementById('cal-wrap');
  wrap.innerHTML='';

  // 依月份分組
  const months={};
  D.forEach(row=>{
    const m=row.date.replace(/\//g,'-').slice(0,7);
    if(!months[m]) months[m]=[];
    months[m].push(row);
  });

  Object.entries(months).forEach(([month, rows])=>{
    const [y,mo]=month.split('-').map(Number); // month已是yyyy-mm格式
    const firstDay=new Date(y,mo-1,1).getDay();
    const daysInMonth=new Date(y,mo,0).getDate();

    // 建立日期→資料的Map
    const dayMap={};
    rows.forEach(r=>{ dayMap[parseInt(r.date.replace(/\//g,'-').slice(8))]=r; });

    const div=document.createElement('div');
    div.className='cal-month';
    const mLabel=`${String(y).slice(2)}/${String(mo).padStart(2,'0')}`;
    div.innerHTML=`<div class="cal-month-title">${mLabel}</div>
<div class="cal-grid">
  ${DOW.map(d=>`<div class="cal-dow">${d}</div>`).join('')}
  ${Array(firstDay).fill('<div class="cal-day empty"></div>').join('')}
  ${Array.from({length:daysInMonth},(_,i)=>{
    const day=i+1;
    const row=dayMap[day];
    const date=new Date(y,mo-1,day);
    const dow=date.getDay();
    const isWeekend=dow===0||dow===6;
    const hasData=row&&row.total>0;
    let badges='';
    if(row){
      AREAS.forEach(a=>{
        const d=row.areas[a];
        if(!d||d.c===0) return;
        const c=d.color;
        badges+=`<div class="cal-badge bg-${c}">
          <div class="cal-badge-dot" style="background:${COLOR_MAP[c]}"></div>
          <span style="color:var(--mu);font-size:10px;flex-shrink:0">${a.slice(0,1)}</span>
          <span class="cal-badge-n" style="color:${COLOR_MAP[c]}">${d.c}</span>
        </div>`;
      });
    }
    return `<div class="cal-day${isWeekend?' weekend':''}${hasData?' has-data':''}">
      <div class="cal-date">${day}</div>
      <div class="cal-badges">${badges}</div>
    </div>`;
  }).join('')}
</div>`;
    wrap.appendChild(div);
  });
}
</script></body></html>"""

    coach_html_out = coach_html.replace('COACH_DATA_PLACEHOLDER', coach_json)
    coach_path = os.path.join(folder, 'dbc_coach_report.html')
    with open(coach_path, 'w', encoding='utf-8') as f:
        f.write(coach_html_out)
    print(f"OK 每日教練需求報表已生成：dbc_coach_report.html")
except Exception as e:
    import traceback
    print(f"教練需求報表生成失敗（不影響主流程）：{e}")
    traceback.print_exc()

# ===== 銷量分析報表 =====
print("\n生成銷量分析報表...")
try:
    import json, re as _re

    _df = df.copy()
    _df['團號_orig'] = _df['團號']

    def _get_area(g):
        if pd.isna(g): return '其他'
        s = str(g)
        if 'ZN' in s: return '野澤'
        if 'ZB' in s: return '斑尾'
        if 'ZY' in s: return '湯澤'
        if 'YS' in s or 'NT' in s: return '龍平'
        if 'BT' in s: return '斑尾'
        if 'N' in s: return '野澤'
        if 'B' in s: return '斑尾'
        return '其他'

    def _get_type(g):
        """判斷純課還是團客"""
        if pd.isna(g): return '團客'
        s = str(g)
        if 'ZN' in s or 'ZB' in s or 'ZY' in s: return '純課'
        return '團客'

    def _norm_level(lv):
        if pd.isna(lv) or str(lv).strip() == '': return ''
        m = _re.search(r'(\d+)', str(lv))
        return f"lv{m.group(1)}" if m else ''

    _df['_area'] = _df['團號_orig'].apply(_get_area)
    _df['_type'] = _df['團號_orig'].apply(_get_type)
    _df['_lv'] = _df['LEVEL'].apply(_norm_level)
    _df['_signup_dt'] = pd.to_datetime(_df['報名日期'], errors='coerce')
    _df['_fee'] = pd.to_numeric(_df['團費'].astype(str).str.replace(',','').str.replace(' ',''), errors='coerce')
    _df['_age'] = pd.to_numeric(_df['年齡'], errors='coerce')

    # 各區 × 純課/團客 人數（從清洗後的all_records_flat取，與課程安排表一致）
    area_type = {}
    for r in all_records_flat:
        item = r.get('項目', '')
        if '野澤' in item: area = '野澤'
        elif '斑尾' in item: area = '斑尾'
        elif '湯澤' in item: area = '湯澤'
        elif '龍平' in item: area = '龍平'
        else: continue
        is_pure = '純課' in item
        if area not in area_type:
            area_type[area] = {'pure': 0, 'group': 0, 'total': 0}
        if is_pure:
            area_type[area]['pure'] += 1
        else:
            area_type[area]['group'] += 1
        area_type[area]['total'] += 1

    # 每日報名（含各區）
    AREAS = ['野澤', '斑尾', '湯澤', '龍平']
    def _get_area2(g):
        if pd.isna(g): return '其他'
        s = str(g)
        if 'ZN' in s or ('N' in s and 'ZB' not in s and 'ZY' not in s and 'NT' not in s): return '野澤'
        if 'ZB' in s or 'BT' in s or 'B' in s: return '斑尾'
        if 'ZY' in s: return '湯澤'
        if 'YS' in s or 'NT' in s: return '龍平'
        return '其他'
    _df['_area2'] = _df['團號_orig'].apply(_get_area2)

    daily_area = _df.groupby([_df['_signup_dt'].dt.date, '_area2']).size().unstack(fill_value=0)
    all_dates = sorted(daily_area.index)
    daily_signup_list = []
    for d in all_dates:
        if not pd.notna(d): continue
        row = {'d': str(d)}
        for a in AREAS:
            row[a] = int(daily_area.loc[d, a]) if a in daily_area.columns else 0
        row['c'] = sum(row.get(a, 0) for a in AREAS)
        daily_signup_list.append(row)

    # 累積報名
    daily_total = _df.groupby(_df['_signup_dt'].dt.date).size().sort_index()
    cum = daily_total.cumsum()
    cum_list = [{'d': str(d), 'c': int(c)} for d, c in cum.items() if pd.notna(d)]

    # 月度報名
    monthly = _df.groupby(_df['_signup_dt'].dt.to_period('M').astype(str)).size()
    monthly_dict = {k: int(v) for k,v in monthly.items() if k not in ('NaT','nan')}

    # 每週報名（含各區）
    _df['_week'] = _df['_signup_dt'].dt.to_period('W').astype(str)
    weekly_area = _df.groupby(['_week', '_area2']).size().unstack(fill_value=0)
    weekly_list = []
    for w in sorted(weekly_area.index):
        if w in ('NaT', 'nan'): continue
        row = {'w': w}
        for a in AREAS:
            row[a] = int(weekly_area.loc[w, a]) if a in weekly_area.columns else 0
        row['c'] = sum(row.get(a, 0) for a in AREAS)
        weekly_list.append(row)

    # 出發月分布（各區）
    _df['_depart_m'] = _df['_signup_dt'].dt.to_period('M').astype(str)  # placeholder
    # 用all_records_flat的出發日期
    depart_data = {}
    for r in all_records_flat:
        dep = r.get('出發日期', '')
        area = r.get('_area_orig', r.get('項目', ''))
        if not dep: continue
        if dep not in depart_data:
            depart_data[dep] = {}
        depart_data[dep][area] = depart_data[dep].get(area, 0) + 1

    # 從all_records_flat重新統計區域
    area_from_records = {}
    for r in all_records_flat:
        item = r.get('項目', '')
        area = '野澤' if '野澤' in item else '斑尾' if '斑尾' in item else '湯澤' if '湯澤' in item else '龍平' if '龍平' in item else '其他'
        area_from_records[area] = area_from_records.get(area, 0) + 1

    # 出發月×區域（從_df）
    area_depart = {}
    for r in all_records_flat:
        dep = r.get('出發日期', '')
        if not dep or len(dep) < 5: continue
        mon = '20' + dep[:2] + '-' + dep[3:5] if len(dep) >= 5 else dep
        item = r.get('項目', '')
        area = '野澤' if '野澤' in item else '斑尾' if '斑尾' in item else '湯澤' if '湯澤' in item else '龍平' if '龍平' in item else '其他'
        if area == '其他': continue
        if mon not in area_depart: area_depart[mon] = {}
        area_depart[mon][area] = area_depart[mon].get(area, 0) + 1

    months_sorted = sorted(area_depart.keys())
    areas_all = ['野澤', '斑尾', '湯澤', '龍平']
    stacked_vals = [[area_depart.get(m, {}).get(a, 0) for a in areas_all] for m in months_sorted]

    # Level（整體，排除空值）
    level_dist = {k: int(v) for k,v in _df['_lv'].value_counts().items() if k}

    # 各區Level - 統一用all_records_flat，與area_type保持一致
    area_level = {'野澤':{},'斑尾':{},'湯澤':{},'龍平':{}}
    for r in all_records_flat:
        item = r.get('項目', '')
        if '野澤' in item: area = '野澤'
        elif '斑尾' in item: area = '斑尾'
        elif '湯澤' in item: area = '湯澤'
        elif '龍平' in item: area = '龍平'
        else: continue
        lv = normalize_level(r.get('LEVEL', ''))
        if not lv: continue
        area_level[area][lv] = area_level[area].get(lv, 0) + 1

    # 驗證：各區level總計應 <= 各區總人數
    for area in ['野澤','斑尾','湯澤','龍平']:
        at = area_type.get(area, {})
        lv_sum = sum(area_level[area].values())
        total = at.get('total', 0)
        if lv_sum > total:
            print(f"  警告：{area} level總計({lv_sum}) > 總人數({total})，請檢查資料")

    # 基礎統計
    s = {
        'pax': len(_df),
        'orders': int(_df['訂單編號'].nunique()),
        'revenue': int(_df['_fee'].sum()),
        'avg': int(_df['_fee'].mean()),
        'female': int((_df['性別']=='女').sum()),
        'male': int((_df['性別']=='男').sum()),
        'female_pct': round(_df[_df['性別']=='女'].shape[0] / max(_df['性別'].isin(['男','女']).sum(), 1) * 100, 1),
        'male_pct': round(_df[_df['性別']=='男'].shape[0] / max(_df['性別'].isin(['男','女']).sum(), 1) * 100, 1),
        'beginner_pct': round(_df[_df['_lv']=='lv0'].shape[0] / max((_df['_lv']!='').sum(), 1) * 100, 1),
        'board_pct': round(_df[_df['滑雪類別']=='單板'].shape[0] / max(_df['滑雪類別'].isin(['單板','雙板']).sum(), 1) * 100, 1),
        'updated': datetime.now().strftime('%Y/%m/%d %H:%M'),
    }

    # 加購
    _with = [c for c in headers if ('續滑' in str(c) or '加滑' in str(c)) and '含' in str(c) and '不含' not in str(c) and '保留' not in str(c)]
    _without = [c for c in headers if ('續滑' in str(c) or '加滑' in str(c)) and '不含' in str(c) and '保留' not in str(c)]
    _excl = [c for c in headers if '專屬' in str(c)]
    _rc = [c for c in headers if '租借' in str(c) and '雪衣' in str(c)]
    _rk = [c for c in headers if '租借' in str(c) and '護膝' in str(c)]
    _rh = [c for c in headers if '租借' in str(c) and '護臀' in str(c)]
    _rg = [c for c in headers if '租借' in str(c) and '手套' in str(c)]

    def _hv(row, cols):
        for c in cols:
            if c in row.index:
                v = row[c]
                if pd.notna(v) and v != 0 and str(v).strip() not in ('', '0'): return True
        return False

    addons = {
        '續滑含教練': int(_df.apply(lambda r: _hv(r, _with), axis=1).sum()),
        '續滑不含': int(_df.apply(lambda r: _hv(r, _without), axis=1).sum()),
        '專屬教練': int(_df.apply(lambda r: _hv(r, _excl), axis=1).sum()),
        '租借衣褲': int(_df.apply(lambda r: _hv(r, _rc), axis=1).sum()),
        '租借護膝': int(_df.apply(lambda r: _hv(r, _rk), axis=1).sum()),
        '租借護臀': int(_df.apply(lambda r: _hv(r, _rh), axis=1).sum()),
        '租借手套': int(_df.apply(lambda r: _hv(r, _rg), axis=1).sum()),
    }

    age_bins = [0,18,25,30,35,40,50,100]
    age_labels = ['18以下','19-25','26-30','31-35','36-40','41-50','51+']
    age_cut = pd.cut(_df['_age'], bins=age_bins, labels=age_labels, right=True)
    age_dist = {str(k): int(v) for k,v in age_cut.value_counts().sort_index().items()}

    rpt = {
        's': s,
        'area': {k: int(v) for k,v in area_from_records.items() if k != '其他'},
        'area_type': area_type,
        'daily': daily_signup_list,
        'weekly': weekly_list,
        'monthly': monthly_dict,
        'cum': cum_list,
        'stacked': {'months': months_sorted, 'areas': areas_all, 'vals': stacked_vals},
        'level': level_dist,
        'area_level': area_level,
        'gender': {str(k): int(v) for k,v in _df['性別'].value_counts().items() if str(k) not in ('nan','')},
        'ski': {str(k): int(v) for k,v in _df['滑雪類別'].value_counts().items() if str(k) not in ('nan','')},
        'age': age_dist,
        'addons': addons,
    }

    rpt_json = json.dumps(rpt, ensure_ascii=False)

    html_template = r"""<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>DBC 銷量分析</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@300;400;500;700&family=DM+Mono:wght@400;500&display=swap');
:root{--bg:#0a0f1e;--s1:#111827;--s2:#1a2235;--bd:#1e2d45;--ac:#3b82f6;--gr:#10b981;--am:#f59e0b;--tx:#e2e8f0;--mu:#64748b;--ny:#3b82f6;--nb:#06b6d4;--nt:#f59e0b;--nl:#8b5cf6;}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--tx);font-family:'Noto Sans TC',sans-serif}
.hd{background:linear-gradient(135deg,#0f172a,#1e3a5f 50%,#0f172a);border-bottom:1px solid var(--bd);padding:14px 20px;position:sticky;top:0;z-index:100}
.hd-in{max-width:1280px;margin:0 auto;display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap}
.logo{display:flex;align-items:center;gap:10px}
.logo-i{width:30px;height:30px;background:var(--ac);border-radius:7px;display:flex;align-items:center;justify-content:center;font-size:15px}
.logo-t h1{font-size:15px;font-weight:700}
.logo-t p{font-size:10px;color:var(--mu);font-family:'DM Mono',monospace}
.badge{background:var(--s2);border:1px solid var(--bd);border-radius:20px;padding:4px 12px;font-size:11px;color:var(--nb);font-family:'DM Mono',monospace}
.main{max-width:1280px;margin:0 auto;padding:18px 16px 48px}
.sec{font-size:11px;font-weight:500;color:var(--mu);text-transform:uppercase;letter-spacing:1px;margin:18px 0 12px;display:flex;align-items:center;gap:8px}
.sec::after{content:'';flex:1;height:1px;background:var(--bd)}
.card{background:var(--s1);border:1px solid var(--bd);border-radius:10px;padding:16px;margin-bottom:14px}
.ct{font-size:12px;font-weight:500;margin-bottom:14px;display:flex;align-items:center;gap:7px;flex-wrap:wrap}
.dot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.cw{position:relative}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.g3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px}

/* ── 總覽大卡 ── */
.overview{display:grid;grid-template-columns:auto 1fr auto;gap:20px;align-items:center}
.ov-total{text-align:center;padding-right:20px;border-right:1px solid var(--bd)}
.ov-total .big{font-size:48px;font-weight:700;font-family:'DM Mono',monospace;color:var(--tx);line-height:1}
.ov-total .lbl{font-size:11px;color:var(--mu);margin-top:4px}
.ov-areas{display:grid;grid-template-columns:repeat(auto-fill,minmax(170px,1fr));gap:10px}
.area-card{background:var(--s2);border:1px solid var(--bd);border-left:3px solid var(--ac);border-radius:8px;padding:12px}
.area-name{font-size:13px;font-weight:600;margin-bottom:6px}
.area-total{font-size:20px;font-weight:700;font-family:'DM Mono',monospace;line-height:1;margin-bottom:6px}
.area-bar{height:5px;background:var(--s1);border-radius:3px;margin-bottom:6px;overflow:hidden;display:flex;gap:1px}
.area-bar-g{height:100%;border-radius:2px}
.area-split{display:flex;justify-content:space-between;font-size:10px;color:var(--mu)}
.area-split span{font-family:'DM Mono',monospace}
.ov-rev{text-align:center;padding-left:20px;border-left:1px solid var(--bd)}
.ov-rev .rev-v{font-size:32px;font-weight:700;font-family:'DM Mono',monospace;color:var(--gr);line-height:1}
.ov-rev .rev-l{font-size:11px;color:var(--mu);margin-top:4px}

/* ── Tabs ── */
.tabs{display:flex;gap:4px;flex-wrap:wrap}
.tab{padding:4px 11px;border-radius:5px;font-size:11px;cursor:pointer;border:1px solid transparent;transition:all .15s;color:var(--mu);background:transparent}
.tab.active{background:var(--ac);border-color:var(--ac);color:#fff;font-weight:500}
.tab:hover:not(.active){border-color:var(--bd);color:var(--tx)}

/* ── 小KPI ── */
.mini-kpi{display:grid;grid-template-columns:repeat(auto-fill,minmax(120px,1fr));gap:10px;margin-bottom:14px}
.mk{background:var(--s2);border:1px solid var(--bd);border-radius:8px;padding:12px;position:relative;overflow:hidden}
.mk::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:var(--mkc,var(--ac))}
.mk-l{font-size:10px;color:var(--mu);text-transform:uppercase;letter-spacing:.6px;margin-bottom:5px}
.mk-v{font-size:20px;font-weight:700;font-family:'DM Mono',monospace;line-height:1}
.mk-s{font-size:10px;color:var(--mu);margin-top:3px}

/* ── 加購 ── */
.addon-g{display:grid;grid-template-columns:repeat(auto-fill,minmax(105px,1fr));gap:8px}
.addon-c{background:var(--s2);border:1px solid var(--bd);border-radius:8px;padding:11px;text-align:center}
.addon-n{font-size:19px;font-weight:700;font-family:'DM Mono',monospace;color:var(--ac)}
.addon-l{font-size:10px;color:var(--mu);margin-top:3px}
.addon-p{font-size:9px;color:var(--mu);font-family:'DM Mono',monospace}

.note{font-size:10px;color:var(--mu);margin-top:8px;line-height:1.5;font-style:italic}

@media(max-width:768px){
  .overview{grid-template-columns:1fr;gap:14px}
  .ov-total{border-right:none;border-bottom:1px solid var(--bd);padding:0 0 14px}
  .ov-rev{border-left:none;border-top:1px solid var(--bd);padding:14px 0 0}
  .g2,.g3{grid-template-columns:1fr}
}
</style></head><body>

<div class="hd"><div class="hd-in">
  <div class="logo"><div class="logo-i">🏔</div><div class="logo-t"><h1>DBC 銷量分析</h1><p id="season-label">SEASON REPORT</p></div></div>
  <div class="badge" id="upd">—</div>
</div></div>

<div class="main">

  <!-- ══ 總覽 ══ -->
  <div class="sec">總覽</div>
  <div class="card">
    <div class="overview">
      <div class="ov-total">
        <div class="big" id="ov-pax">—</div>
        <div class="lbl">總報名人次</div>
      </div>
      <div class="ov-areas" id="ov-areas"></div>
      <div class="ov-rev">
        <div class="rev-v" id="ov-rev">—</div>
        <div class="rev-l">總營收 NT$</div>
      </div>
    </div>
  </div>

  <!-- ══ 累積報名 ══ -->
  <div class="sec">累積報名人次</div>
  <div class="card">
    <div class="ct">
      <span class="dot" style="background:#10b981"></span>累積報名曲線
      <div class="tabs" style="margin-left:auto">
        <button class="tab active" onclick="setCum('all',this)">全期</button>
        <button class="tab" onclick="setCum('2m',this)">近2月</button>
        <button class="tab" onclick="setCum('1m',this)">近1月</button>
      </div>
    </div>
    <div class="cw"><canvas id="cumC" height="90"></canvas></div>
  </div>

  <!-- ══ 每日報名 ══ -->
  <div class="sec">每日新增報名</div>
  <div class="card">
    <div class="ct">
      <span class="dot" style="background:#f59e0b"></span>每日新增
      <div class="tabs" style="margin-left:auto" id="daily-tabs">
        <button class="tab active" onclick="setDaily('all','all',this)">全部</button>
        <button class="tab" onclick="setDaily('all','野澤',this)">野澤</button>
        <button class="tab" onclick="setDaily('all','斑尾',this)">斑尾</button>
        <button class="tab" onclick="setDaily('all','湯澤',this)">湯澤</button>
        <button class="tab" onclick="setDaily('all','龍平',this)">龍平</button>
        <button class="tab" onclick="setDaily('3m','all',this)">近3月</button>
      </div>
    </div>
    <div class="cw"><canvas id="dailyC" height="140"></canvas></div>
    <div class="note">峰值通常對應KOL發布或廣告上架時間點。可切換各區單獨觀察。</div>
  </div>

  <!-- ══ 每週報名 ══ -->
  <div class="sec">每週新增報名</div>
  <div class="card">
    <div class="ct">
      <span class="dot" style="background:#8b5cf6"></span>每週新增
      <div class="tabs" style="margin-left:auto" id="weekly-tabs">
        <button class="tab active" onclick="setWeekly('all',this)">全部</button>
        <button class="tab" onclick="setWeekly('野澤',this)">野澤</button>
        <button class="tab" onclick="setWeekly('斑尾',this)">斑尾</button>
        <button class="tab" onclick="setWeekly('湯澤',this)">湯澤</button>
        <button class="tab" onclick="setWeekly('龍平',this)">龍平</button>
      </div>
    </div>
    <div class="cw"><canvas id="weeklyC" height="140"></canvas></div>
  </div>

  <!-- ══ 出發日期分析 ══ -->
  <div class="sec">出發日期分析</div>
  <div class="g2">
    <div class="card" style="margin-bottom:0">
      <div class="ct"><span class="dot" style="background:#06b6d4"></span>各區出發人次</div>
      <div id="area-bars"></div>
    </div>
    <div class="card" style="margin-bottom:0">
      <div class="ct"><span class="dot" style="background:#3b82f6"></span>出發月份 × 各區</div>
      <div class="cw"><canvas id="stackC" height="200"></canvas></div>
    </div>
  </div>

  <!-- ══ 客群分析 ══ -->
  <div class="sec" style="margin-top:14px">客群分析</div>
  <div class="mini-kpi" id="demo-kpi"></div>
  <div class="g3">
    <div class="card" style="margin-bottom:0">
      <div class="ct"><span class="dot" style="background:#06b6d4"></span>Level 分布</div>
      <div class="cw"><canvas id="lvC" height="200"></canvas></div>
    </div>
    <div class="card" style="margin-bottom:0">
      <div class="ct"><span class="dot" style="background:#3b82f6"></span>年齡分布</div>
      <div class="cw"><canvas id="ageC" height="200"></canvas></div>
    </div>
    <div class="card" style="margin-bottom:0">
      <div class="ct"><span class="dot" style="background:#8b5cf6"></span>各區Level
        <div class="tabs" style="margin-left:auto">
          <button class="tab active" onclick="setAreaLv('野澤',this)">野澤</button>
          <button class="tab" onclick="setAreaLv('斑尾',this)">斑尾</button>
          <button class="tab" onclick="setAreaLv('湯澤',this)">湯澤</button>
          <button class="tab" onclick="setAreaLv('龍平',this)">龍平</button>
        </div>
      </div>
      <div class="cw"><canvas id="areaLvC" height="170"></canvas></div>
    </div>
  </div>

  <!-- ══ 加購 ══ -->
  <div class="sec" style="margin-top:14px">加購 & 租借</div>
  <div class="card">
    <div class="ct"><span class="dot" style="background:#10b981"></span>加購項目人次</div>
    <div class="addon-g" id="addon-g"></div>
  </div>

</div>
<script>
const D=REPORT_DATA_PLACEHOLDER;
const AC={'野澤':'#ffffff','斑尾':'#7ec8f5','湯澤':'#ffb347','龍平':'#c8864a'};
const LC={'lv0':'#64748b','lv1':'#3b82f6','lv2':'#06b6d4','lv3':'#f59e0b','lv4':'#ef4444','lv5':'#8b5cf6'};
const LL={'lv0':'lv0 初學','lv1':'lv1 入門','lv2':'lv2 初中','lv3':'lv3 中級','lv4':'lv4 高級','lv5':'lv5 專業'};
const AREAS=['野澤','斑尾','湯澤','龍平'];
function co(h,a){const r=parseInt(h.slice(1,3),16),g=parseInt(h.slice(3,5),16),b=parseInt(h.slice(5,7),16);return `rgba(${r},${g},${b},${a})`}
const fmt=n=>n>=100000000?(n/100000000).toFixed(2)+'億':n>=10000?(n/10000).toFixed(1)+'萬':n.toLocaleString();
const CD={responsive:true,plugins:{legend:{display:false},tooltip:{backgroundColor:'#1a2235',borderColor:'#1e2d45',borderWidth:1,titleColor:'#e2e8f0',bodyColor:'#94a3b8',padding:10,cornerRadius:8}},scales:{x:{grid:{color:'#1e2d45'},ticks:{color:'#64748b',font:{size:10},maxTicksLimit:14}},y:{grid:{color:'#1e2d45'},ticks:{color:'#64748b',font:{size:10}}}}};

// ── 總覽 ──
document.getElementById('ov-pax').textContent=D.s.pax.toLocaleString();
document.getElementById('ov-rev').textContent=fmt(D.s.revenue);
document.getElementById('upd').textContent='更新：'+D.s.updated;
const og=document.getElementById('ov-areas');
const tot=D.s.pax;
AREAS.forEach(a=>{
  const at=D.area_type[a]||{pure:0,group:0,total:0};
  if(!at.total)return;
  const c=AC[a];
  const gPct=at.total>0?(at.group/at.total*100).toFixed(0):0;
  const pPct=at.total>0?(at.pure/at.total*100).toFixed(0):0;
  const aPct=(at.total/tot*100).toFixed(1);
  og.innerHTML+=`<div class="area-card" style="border-left-color:${c}">
    <div class="area-name" style="color:${c}">${a} <span style="font-size:10px;color:var(--mu);font-weight:400">${aPct}%</span></div>
    <div class="area-total">${at.total.toLocaleString()}<span style="font-size:11px;color:var(--mu);font-weight:400"> 人</span></div>
    <div class="area-bar"><div class="area-bar-g" style="flex:${at.group};background:${c};opacity:.9"></div><div class="area-bar-g" style="flex:${at.pure};background:${c};opacity:.35"></div></div>
    <div class="area-split"><span style="color:${c}">團${at.group.toLocaleString()} (${gPct}%)</span><span style="opacity:.6">純課${at.pure.toLocaleString()} (${pPct}%)</span></div>
  </div>`;
});

// ── 累積 ──
let cumChart=null;
function setCum(range,btn){
  document.querySelectorAll('#cumC').forEach(()=>{});
  document.querySelectorAll('.tab').forEach(t=>{if(t.onclick&&t.onclick.toString().includes('setCum'))t.classList.remove('active')});
  btn.classList.add('active');
  let pts=D.cum;
  if(range!=='all'){const days=range==='1m'?30:60;const cut=new Date(pts[pts.length-1].d);cut.setDate(cut.getDate()-days);pts=pts.filter(p=>new Date(p.d)>=cut);}
  if(cumChart)cumChart.destroy();
  cumChart=new Chart(document.getElementById('cumC'),{type:'line',data:{labels:pts.map(p=>p.d.slice(5)),datasets:[{data:pts.map(p=>p.c),borderColor:'#10b981',backgroundColor:co('#10b981',.1),fill:true,tension:.4,borderWidth:2,pointRadius:0}]},options:{...CD,scales:{x:{...CD.scales.x,ticks:{...CD.scales.x.ticks,maxTicksLimit:16}},y:{...CD.scales.y}}}});
}
setCum('all',document.querySelector('.tab'));

// ── 每日 ──
let dailyChart=null;
function setDaily(range,area,btn){
  document.querySelectorAll('#daily-tabs .tab').forEach(t=>t.classList.remove('active'));
  btn.classList.add('active');
  let pts=D.daily;
  if(range==='3m'){const cut=new Date(pts[pts.length-1].d);cut.setDate(cut.getDate()-90);pts=pts.filter(p=>new Date(p.d)>=cut);}
  if(dailyChart)dailyChart.destroy();
  const labels=pts.map(p=>p.d.slice(5));
  let datasets;
  if(area==='all'){
    datasets=AREAS.map(a=>({label:a,data:pts.map(p=>p[a]||0),backgroundColor:co(AC[a],.8),borderRadius:1,borderSkipped:false}));
  } else {
    datasets=[{label:area,data:pts.map(p=>p[area]||0),backgroundColor:co(AC[area],.8),borderRadius:2,borderSkipped:false}];
  }
  const stacked=area==='all';
  dailyChart=new Chart(document.getElementById('dailyC'),{type:'bar',data:{labels,datasets},options:{...CD,plugins:{...CD.plugins,legend:{display:stacked,position:'bottom',labels:{color:'#94a3b8',font:{size:10},boxWidth:10,padding:10}}},scales:{x:{...CD.scales.x,stacked,ticks:{...CD.scales.x.ticks,maxTicksLimit:18}},y:{...CD.scales.y,stacked}}}});
}
setDaily('all','all',document.querySelectorAll('#daily-tabs .tab')[0]);

// ── 每週 ──
let weeklyChart=null;
function setWeekly(area,btn){
  document.querySelectorAll('#weekly-tabs .tab').forEach(t=>t.classList.remove('active'));
  btn.classList.add('active');
  const pts=D.weekly;
  const labels=pts.map(p=>p.w.slice(5,12));
  let datasets;
  if(area==='all'){
    datasets=AREAS.map(a=>({label:a,data:pts.map(p=>p[a]||0),backgroundColor:co(AC[a],.8),borderRadius:1,borderSkipped:false}));
  } else {
    datasets=[{label:area,data:pts.map(p=>p[area]||0),backgroundColor:co(AC[area],.8),borderRadius:2,borderSkipped:false}];
  }
  const stacked=area==='all';
  if(weeklyChart)weeklyChart.destroy();
  weeklyChart=new Chart(document.getElementById('weeklyC'),{type:'bar',data:{labels,datasets},options:{...CD,plugins:{...CD.plugins,legend:{display:stacked,position:'bottom',labels:{color:'#94a3b8',font:{size:10},boxWidth:10,padding:10}}},scales:{x:{...CD.scales.x,stacked,ticks:{...CD.scales.x.ticks,maxTicksLimit:16}},y:{...CD.scales.y,stacked}}}});
}
setWeekly('all',document.querySelectorAll('#weekly-tabs .tab')[0]);

// ── 出發分布 ──
const areaBars=document.getElementById('area-bars');
AREAS.forEach(a=>{
  const at=D.area_type[a]||{total:0};
  if(!at.total)return;
  const pct=(at.total/tot*100).toFixed(1);
  const c=AC[a];
  areaBars.innerHTML+=`<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px"><div style="font-size:11px;color:var(--mu);min-width:45px;text-align:right">${a}</div><div style="flex:1;height:22px;background:var(--s2);border-radius:5px;overflow:hidden"><div style="width:${pct}%;height:100%;background:${c};opacity:.8;display:flex;align-items:center;padding-left:8px;font-size:10px;font-family:'DM Mono',monospace;color:#fff">${at.total.toLocaleString()}</div></div><div style="font-size:10px;color:var(--mu);font-family:'DM Mono',monospace;min-width:36px">${pct}%</div></div>`;
});

const sMonths=D.stacked.months.map(m=>m.replace('2025-','').replace('2026-','').replace('2027-',''));
new Chart(document.getElementById('stackC'),{type:'bar',data:{labels:sMonths,datasets:D.stacked.areas.map((a,ai)=>({label:a,data:D.stacked.vals.map(r=>r[ai]),backgroundColor:co(AC[a]||'#64748b',.8),borderRadius:2,borderSkipped:false}))},options:{...CD,plugins:{...CD.plugins,legend:{display:true,position:'bottom',labels:{color:'#94a3b8',font:{size:10},boxWidth:10,padding:10}}},scales:{x:{...CD.scales.x,stacked:true},y:{...CD.scales.y,stacked:true}}}});

// ── 客群mini KPI ──
const dk=document.getElementById('demo-kpi');
[
  {l:'女性',v:D.s.female,s:D.s.female_pct+'%',c:'#ec4899'},
  {l:'男性',v:D.s.male,s:D.s.male_pct+'%',c:'#3b82f6'},
  {l:'單板',v:D.ski['單板']||0,s:D.s.board_pct+'%',c:'#06b6d4'},
  {l:'雙板',v:D.ski['雙板']||0,s:(100-D.s.board_pct).toFixed(1)+'%',c:'#f59e0b'},
  {l:'初學(lv0)',v:D.level['lv0']||0,s:D.s.beginner_pct+'%',c:'#8b5cf6'},
].forEach(item=>{
  dk.innerHTML+=`<div class="mk" style="--mkc:${item.c}"><div class="mk-l">${item.l}</div><div class="mk-v" style="color:${item.c}">${item.v.toLocaleString()}</div><div class="mk-s">${item.s}</div></div>`;
});

// ── Level整體 ──
const lvK=Object.keys(D.level),lvV=Object.values(D.level);
new Chart(document.getElementById('lvC'),{type:'bar',data:{labels:lvK.map(k=>LL[k]||k),datasets:[{data:lvV,backgroundColor:lvK.map(k=>co(LC[k]||'#64748b',.8)),borderRadius:3,borderSkipped:false}]},options:{...CD}});

// ── 年齡 ──
const ageK=Object.keys(D.age),ageV=Object.values(D.age);
new Chart(document.getElementById('ageC'),{type:'bar',data:{labels:ageK,datasets:[{data:ageV,backgroundColor:ageV.map((_,i)=>co('#f59e0b',.4+i/ageK.length*.6)),borderRadius:3,borderSkipped:false}]},options:{...CD,indexAxis:'y',scales:{x:{...CD.scales.x},y:{...CD.scales.y}}}});

// ── 各區Level ──
let aLvChart=null;
function setAreaLv(area,btn){
  document.querySelectorAll('.tab').forEach(t=>{if(t.onclick&&t.onclick.toString().includes('setAreaLv'))t.classList.remove('active')});
  btn.classList.add('active');
  const d=D.area_level[area]||{};
  const keys=['lv0','lv1','lv2','lv3','lv4'];
  if(aLvChart)aLvChart.destroy();
  aLvChart=new Chart(document.getElementById('areaLvC'),{type:'bar',data:{labels:keys.map(k=>LL[k]||k),datasets:[{data:keys.map(k=>d[k]||0),backgroundColor:keys.map(k=>co(LC[k]||'#64748b',.8)),borderRadius:3,borderSkipped:false}]},options:{...CD}});
}
setAreaLv('野澤',document.querySelectorAll('.tab')[document.querySelectorAll('.tab').length-4]);

// ── 加購 ──
const ag=document.getElementById('addon-g');
Object.entries(D.addons).forEach(([name,val])=>{
  const pct=(val/D.s.pax*100).toFixed(1);
  ag.innerHTML+=`<div class="addon-c"><div class="addon-n">${val.toLocaleString()}</div><div class="addon-l">${name}</div><div class="addon-p">${pct}% 人次</div></div>`;
});
</script></body></html>"""

    html_out = html_template.replace('REPORT_DATA_PLACEHOLDER', rpt_json)
    report_path = os.path.join(folder, 'dbc_sales_report.html')
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(html_out)
    print(f"OK 銷量分析報表已生成：dbc_sales_report.html")
except Exception as e:
    print(f"銷量報表生成失敗（不影響主流程）：{e}")

print(f"\n{'='*50}")
print(f"全部完成！{datetime.now().strftime('%Y/%m/%d %H:%M')}")
print(f"{'='*50}")
